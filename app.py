#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ==========================================================
#  VERSION 1.7 - 2 de septiembre de 2026, 6:20 pm (centro MX)
#  v1.1 columnas departamento/mundo/tipo_interno en CREATE TABLE
#  v1.2 modales faltantes en portal.html
#  v1.3 auditoria: 4 pantallas muertas por botones inexistentes
#  v1.4 CORRECCION: el blindaje de la 1.3 rompio la sintaxis de
#       app.js y validacion.js. Rehecho con metodo seguro.
#       Probado en navegador real: 0 errores de JavaScript.
#  v1.5 NUEVO: consulta de dias pasados (/historico). Ver como estaba
#       la obra en una fecha, comparar dos fechas, por responsable y
#       por bloque, en los dos mundos. Reconstruye el pasado desde el
#       historial cuando no hay foto. Foto semanal AUTOMATICA.
#  v1.6 LISTA PARA INTERNET: base en disco persistente (DATA_DIR),
#       puerto y clave desde el servidor, cookies seguras en HTTPS,
#       arranque compatible con gunicorn. Ver GUIA_PUBLICAR_EN_INTERNET.md
#  v1.7 SIEMBRA AUTOMATICA: en un servidor nuevo con disco vacio, copia
#       semilla/obra_inicial.db (955 partidas) la primera vez. Ya no hay
#       que subir la base a mano.
#       VISTA DE CELULAR corregida: los botones del encabezado ya no se
#       salen de la pantalla; tablero, filtros y tabla adaptados.
# ==========================================================
"""
PLATAFORMA DE CONTROL DE OBRA — Unidad Quirúrgica HAP
Servidor local. Base de datos SQLite en ./data/obra.db
Entrega meta: 31 de octubre de 2026.

Uso:
    python app.py
Luego abre en tu navegador:  http://localhost:5000
"""

import os
import sqlite3
import datetime
import json
import hashlib
import secrets
from functools import wraps
from flask import Flask, request, jsonify, render_template, g, send_file, Response, session, redirect

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# En internet (Render/Railway) la base debe vivir en un DISCO PERSISTENTE,
# si no se borra en cada actualizacion. Se define con la variable DATA_DIR.
DATA_DIR = os.environ.get("DATA_DIR") or os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
DB_PATH = os.path.join(DATA_DIR, "obra.db")
FECHA_ENTREGA = "2026-10-31"

app = Flask(__name__)
# clave para las sesiones; se guarda en data/ para que no cambie entre reinicios
_secret_file = os.path.join(DATA_DIR, ".secret")
def _cargar_secret():
    try:
        if os.path.exists(_secret_file):
            return open(_secret_file).read().strip()
        os.makedirs(os.path.dirname(_secret_file), exist_ok=True)
        s = secrets.token_hex(32)
        open(_secret_file, "w").write(s)
        return s
    except Exception:
        return "hap-clave-temporal-cambiar"
app.secret_key = os.environ.get("SECRET_KEY") or _cargar_secret()

# --- Seguridad de la sesion cuando corre en internet (HTTPS) ---
EN_INTERNET = bool(os.environ.get("EN_INTERNET"))
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=EN_INTERNET,
    PERMANENT_SESSION_LIFETIME=datetime.timedelta(hours=12),
)


SAL = os.environ.get("SAL_CLAVES", "")

def hash_clave(txt):
    """Con SAL_CLAVES definida usa sal; sin ella mantiene el formato
    anterior para no invalidar las contrasenas ya creadas."""
    base = (txt or "")
    if SAL:
        return hashlib.sha256((SAL + base).encode()).hexdigest()
    return hashlib.sha256(base.encode()).hexdigest()


def usuario_actual():
    return session.get("usuario"), session.get("rol"), session.get("proveedor")


def col_duenio():
    """Devuelve ('departamento', mundo) o ('proveedor', mundo) según el mundo del usuario."""
    mundo = session.get("mundo", "obra")
    col = "departamento" if mundo == "interno" else "proveedor"
    return col, mundo


def requiere_login(fn):
    @wraps(fn)
    def envoltura(*a, **k):
        if "usuario" not in session:
            return jsonify({"error": "no_login"}), 401
        return fn(*a, **k)
    return envoltura


def requiere_admin(fn):
    @wraps(fn)
    def envoltura(*a, **k):
        if session.get("rol") != "admin":
            return jsonify({"error": "solo_admin"}), 403
        return fn(*a, **k)
    return envoltura


# Roles que pueden GESTIONAR la obra (validar, asignar, editar, dependencias):
# el admin y los dos supervisores. NO incluye tocar usuarios, respaldos ni borrar.
ROLES_GESTORES = ("admin", "supervisor_obra", "supervisor_depto")

def es_gestor(rol):
    return rol in ROLES_GESTORES

def requiere_gestor(fn):
    """Deja pasar a admin y supervisores. Para validar/asignar/editar/dependencias."""
    @wraps(fn)
    def envoltura(*a, **k):
        if session.get("rol") not in ROLES_GESTORES:
            return jsonify({"error": "solo_gestores"}), 403
        return fn(*a, **k)
    return envoltura

# ----------------------------------------------------------------------------
# Base de datos
# ----------------------------------------------------------------------------
import unicodedata


def sin_acentos(txt):
    """Quita acentos y pasa a minúsculas, para buscar sin que estorbe la tilde."""
    if txt is None:
        return ""
    t = unicodedata.normalize("NFD", str(txt))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return t.lower()


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=10.0)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode = WAL")
        g.db.execute("PRAGMA busy_timeout = 5000")
        g.db.execute("PRAGMA foreign_keys = ON")
        # función propia para buscar ignorando acentos y mayúsculas
        g.db.create_function("sinac", 1, sin_acentos)
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def generar_respaldo_bd():
    """Genera una copia de seguridad timestamped en data/backups/ (mantiene los últimos 30)."""
    try:
        backup_dir = os.path.join(BASE_DIR, "data", "backups")
        os.makedirs(backup_dir, exist_ok=True)
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(backup_dir, f"obra_backup_{ts}.db")
        if os.path.exists(DB_PATH):
            src_con = sqlite3.connect(DB_PATH)
            dst_con = sqlite3.connect(dest)
            src_con.backup(dst_con)
            dst_con.close()
            src_con.close()
            # Limpiar respaldos viejos (mantener los últimos 30)
            archivos = sorted([os.path.join(backup_dir, f) for f in os.listdir(backup_dir) if f.startswith("obra_backup_") and f.endswith(".db")])
            if len(archivos) > 30:
                for f in archivos[:-30]:
                    try: os.remove(f)
                    except Exception: pass
            return dest
    except Exception as e:
        print(f"Error generando respaldo automático: {e}")
    return None


def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH, timeout=10.0)
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA busy_timeout = 5000")
    con.executescript(
        """
    CREATE TABLE IF NOT EXISTS actividades (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo         TEXT,
        bloque         TEXT,
        area           TEXT,
        giro           TEXT,
        proveedor      TEXT,
        departamento   TEXT,
        mundo          TEXT DEFAULT 'obra',
        tipo_interno   TEXT,
        partida        TEXT,
        tipo           TEXT,
        tipo_partida   TEXT DEFAULT 'Construcción',
        aplica         TEXT DEFAULT 'SÍ',
        avance         INTEGER DEFAULT 0,
        f_inicio       TEXT,
        f_fin          TEXT,
        duracion_dias  INTEGER,
        estatus        TEXT DEFAULT 'Pendiente',
        depende_de     INTEGER,
        definido       TEXT DEFAULT 'NO',
        causa_retraso  TEXT,
        nota_proveedor TEXT,
        causa_por      TEXT,
        causa_fecha    TEXT,
        -- Portal de proveedores: origen y estado de validación
        origen         TEXT DEFAULT 'oficial',   -- 'oficial' (la diste de alta tú) | 'propuesta' (la agregó el proveedor)
        estado_val     TEXT DEFAULT 'validado',  -- 'validado' | 'propuesta' (espera tu firma) | 'rechazada'
        avance_decl    INTEGER,                  -- lo que el proveedor DICE que lleva (declarado)
        avance_decl_por TEXT,                     -- quién lo declaró
        avance_decl_fecha TEXT,                   -- cuándo lo declaró
        definido_por   TEXT,                      -- 'plano' | 'adicional' | 'comentario' (para que el proveedor se proteja)
        creado_por     TEXT,                      -- usuario que la creó (si es propuesta)
        reconocida     TEXT DEFAULT 'NO',         -- 'SÍ' cuando el proveedor acepta que es su trabajo
        reconocida_por TEXT,
        reconocida_fecha TEXT,
        no_reconocida_nota TEXT,                   -- si el proveedor dice "esto no es mío", su comentario
        fuera_zona     INTEGER DEFAULT 0,
        rechazo_motivo TEXT,
        rechazado_por  TEXT,
        rechazado_fecha TEXT,
        avance_decl_rechazado INTEGER DEFAULT 0,
        notas          TEXT,
        actualizado    TEXT,
        FOREIGN KEY (depende_de) REFERENCES actividades(id) ON DELETE SET NULL
    );

    -- Modelo de Dependencias por área (auto-liberación y forzado con historial)
    CREATE TABLE IF NOT EXISTS dependencias (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        area               TEXT NOT NULL,
        tipo_sucesor       TEXT NOT NULL,
        tipos_predecesores TEXT NOT NULL,
        umbral             INTEGER DEFAULT 100,
        permite_paralelo   INTEGER DEFAULT 0,
        liberacion_forzada INTEGER DEFAULT 0,
        forzada_por        TEXT,
        forzada_fecha      TEXT,
        forzada_nota       TEXT,
        creado             TEXT
    );

    -- Usuarios del portal de proveedores (uno o varios por empresa)
    CREATE TABLE IF NOT EXISTS usuarios (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        usuario   TEXT UNIQUE,     -- lo que teclean para entrar
        clave     TEXT,            -- contraseña simple (se guarda con hash)
        proveedor TEXT,            -- a qué proveedor/departamento pertenece (ve solo lo suyo)
        rol       TEXT DEFAULT 'proveedor',  -- 'proveedor' | 'admin'
        mundo     TEXT DEFAULT 'obra',       -- 'obra' (proveedor externo) | 'interno' (departamento HAP)
        clave_cambiada INTEGER DEFAULT 0,     -- 1 si el usuario ya cambió la contraseña que le diste
        creado    TEXT
    );

    -- Bitácora de avisos para el panel de validación del admin
    CREATE TABLE IF NOT EXISTS avisos (
        id        INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha     TEXT,
        proveedor TEXT,
        usuario   TEXT,
        tipo      TEXT,            -- 'propuesta' | 'avance'
        detalle   TEXT,
        visto     INTEGER DEFAULT 0
    );

    -- Catálogo de causas de retraso (7 base + las que agregue el usuario)
    CREATE TABLE IF NOT EXISTS causas (
        id     INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE,
        base   INTEGER DEFAULT 0,
        mundo  TEXT DEFAULT 'obra'
    );

    -- Fotografías semanales de avance por partida (para comparar semana vs semana)
    CREATE TABLE IF NOT EXISTS snapshots (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        semana       TEXT,      -- etiqueta 'AAAA-Www' (año-semana ISO)
        fecha        TEXT,      -- fecha en que se tomó la foto
        actividad_id INTEGER,
        avance       INTEGER
    );

    CREATE TABLE IF NOT EXISTS historial (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        actividad_id INTEGER,
        campo        TEXT,
        valor_antes  TEXT,
        valor_despues TEXT,
        fecha        TEXT,
        quien        TEXT
    );

    -- Catálogo de áreas/departamentos que intervienen para dejar un espacio funcional
    CREATE TABLE IF NOT EXISTS tipos_internos (
        id     INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE
    );

    CREATE TABLE IF NOT EXISTS involucrados (
        id     INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE
    );

    -- Qué involucrados le tocan a cada área (para dejarla funcional)
    CREATE TABLE IF NOT EXISTS area_involucrados (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        area         TEXT,
        involucrado  TEXT,
        UNIQUE(area, involucrado)
    );

    CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY,
        valor TEXT
    );

    -- Catálogo de proveedores con tipo (interno/externo) y de qué se encarga
    CREATE TABLE IF NOT EXISTS proveedores (
        id       INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre   TEXT UNIQUE,
        tipo     TEXT DEFAULT 'Externo',
        funcion  TEXT
    );

    -- Catálogo simple para bloques, áreas y giros (nombre + nota de para qué es)
    CREATE TABLE IF NOT EXISTS catalogo (
        id     INTEGER PRIMARY KEY AUTOINCREMENT,
        clase  TEXT,      -- 'bloque' | 'area' | 'giro'
        nombre TEXT,
        nota   TEXT,
        UNIQUE(clase, nombre)
    );
    """
    )
    # --- Migración: agrega columnas nuevas si la BD ya existía sin ellas ---
    cols = [r[1] for r in con.execute("PRAGMA table_info(actividades)").fetchall()]
    if "tipo_partida" not in cols:
        con.execute("ALTER TABLE actividades ADD COLUMN tipo_partida TEXT DEFAULT 'Construcción'")
        con.execute("UPDATE actividades SET tipo_partida='Construcción' WHERE tipo_partida IS NULL")
    if "definido" not in cols:
        con.execute("ALTER TABLE actividades ADD COLUMN definido TEXT DEFAULT 'NO'")
        con.execute("UPDATE actividades SET definido='NO' WHERE definido IS NULL")
    for c in ("causa_retraso", "nota_proveedor", "causa_por", "causa_fecha"):
        if c not in cols:
            con.execute(f"ALTER TABLE actividades ADD COLUMN {c} TEXT")
    # Portal de proveedores y rechazos
    portal_cols = {
        "origen": "TEXT DEFAULT 'oficial'",
        "estado_val": "TEXT DEFAULT 'validado'",
        "avance_decl": "INTEGER",
        "avance_decl_por": "TEXT",
        "avance_decl_fecha": "TEXT",
        "definido_por": "TEXT",
        "creado_por": "TEXT",
        "reconocida": "TEXT DEFAULT 'NO'",
        "reconocida_por": "TEXT",
        "reconocida_fecha": "TEXT",
        "no_reconocida_nota": "TEXT",
        "mundo": "TEXT DEFAULT 'obra'",
        "departamento": "TEXT",
        "tipo_interno": "TEXT",
        "fuera_zona": "INTEGER DEFAULT 0",
        "rechazo_motivo": "TEXT",
        "rechazado_por": "TEXT",
        "rechazado_fecha": "TEXT",
        "avance_decl_rechazado": "INTEGER DEFAULT 0",
    }
    for c, tipo in portal_cols.items():
        if c not in cols:
            con.execute(f"ALTER TABLE actividades ADD COLUMN {c} {tipo}")
    con.execute("UPDATE actividades SET mundo='obra' WHERE mundo IS NULL")
    # las partidas que ya existían son oficiales y validadas
    con.execute("UPDATE actividades SET origen='oficial' WHERE origen IS NULL")
    con.execute("UPDATE actividades SET estado_val='validado' WHERE estado_val IS NULL")
    # todas arrancan SIN reconocer: el proveedor tiene que aceptarlas sí o sí
    con.execute("UPDATE actividades SET reconocida='NO' WHERE reconocida IS NULL")

    # historial: columna 'quien' si falta
    hcols = [r[1] for r in con.execute("PRAGMA table_info(historial)").fetchall()]
    if "quien" not in hcols:
        con.execute("ALTER TABLE historial ADD COLUMN quien TEXT")

    # usuarios: columnas mundo y clave_cambiada si faltan
    ucols = [r[1] for r in con.execute("PRAGMA table_info(usuarios)").fetchall()]
    if "mundo" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN mundo TEXT DEFAULT 'obra'")
        con.execute("UPDATE usuarios SET mundo='obra' WHERE mundo IS NULL")
    if "clave_cambiada" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN clave_cambiada INTEGER DEFAULT 0")
    if "activo" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN activo INTEGER DEFAULT 1")
        con.execute("UPDATE usuarios SET activo=1 WHERE activo IS NULL")
    if "num_logins" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN num_logins INTEGER DEFAULT 0")
    if "ultimo_login" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN ultimo_login TEXT")
    if "telefono" not in ucols:
        con.execute("ALTER TABLE usuarios ADD COLUMN telefono TEXT")

    # Expediente de proveedores: columnas de contacto
    pcols = [r[1] for r in con.execute("PRAGMA table_info(proveedores)").fetchall()]
    for c in ("empresa", "contacto", "telefono", "correo", "notas"):
        if c not in pcols:
            con.execute(f"ALTER TABLE proveedores ADD COLUMN {c} TEXT")

    # Sembrar las 7 causas base (solo si la tabla está vacía)
    n_causas = con.execute("SELECT COUNT(*) FROM causas").fetchone()[0]
    if n_causas == 0:
        base = [
            "Definición pendiente de proyecto",
            "Disponibilidad de personal",
            "Suministro o entrega de material",
            "Dependencia de otro proveedor",
            "Cambio solicitado por el hospital",
            "Acceso o liberación del área",
            "Condición no prevista en sitio",
        ]
        for nombre in base:
            con.execute("INSERT OR IGNORE INTO causas (nombre,base,mundo) VALUES (?,1,'obra')", (nombre,))

    # migración: columna mundo en causas si falta
    ccols = [r[1] for r in con.execute("PRAGMA table_info(causas)").fetchall()]
    if "mundo" not in ccols:
        con.execute("ALTER TABLE causas ADD COLUMN mundo TEXT DEFAULT 'obra'")
        con.execute("UPDATE causas SET mundo='obra' WHERE mundo IS NULL")

    # Causas propias del mundo interno (departamentos HAP) — distintas a las de obra
    n_causas_int = con.execute("SELECT COUNT(*) FROM causas WHERE mundo='interno'").fetchone()[0]
    if n_causas_int == 0:
        base_int = [
            "Compras no ha surtido el material",
            "Compras dio fecha de entrega lejana",
            "Falta autorización de la dirección",
            "Presupuesto no liberado",
            "Depende de que termine la obra en el área",
            "Falta definición técnica del equipo",
            "Personal del departamento ocupado en otra prioridad",
            "Equipo en revisión o garantía",
        ]
        for nombre in base_int:
            con.execute("INSERT OR IGNORE INTO causas (nombre,base,mundo) VALUES (?,1,'interno')", (nombre,))

    # Catálogo de involucrados (áreas del hospital que intervienen para dejar funcional un espacio)
    n_inv = con.execute("SELECT COUNT(*) FROM involucrados").fetchone()[0]
    if n_inv == 0:
        base_inv = ["Sistemas", "Intendencia", "Biomédica", "Publicidad", "Compras",
                    "Finanzas", "Contabilidad", "Mantenimiento", "Enfermería", "Gerencia médica"]
        for nombre in base_inv:
            con.execute("INSERT OR IGNORE INTO involucrados (nombre) VALUES (?)", (nombre,))

    # Tipos de tarea del mundo interno (editables por el admin)
    n_ti = con.execute("SELECT COUNT(*) FROM tipos_internos").fetchone()[0]
    if n_ti == 0:
        for nombre in ["Instalación", "Adecuación del área", "Solicitar a compras", "Puesta en marcha / prueba"]:
            con.execute("INSERT OR IGNORE INTO tipos_internos (nombre) VALUES (?)", (nombre,))

    # Usuario admin inicial (Jimmy). Contraseña por defecto: cambiar después.
    import hashlib
    n_users = con.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0]
    if n_users == 0:
        h = hashlib.sha256("HAP2026".encode()).hexdigest()
        con.execute(
            "INSERT INTO usuarios (usuario,clave,proveedor,rol,creado) VALUES (?,?,?,?,?)",
            ("admin", h, None, "admin", datetime.datetime.now().isoformat(timespec="seconds")))
    con.commit()
    con.close()
    generar_respaldo_bd()


# ----------------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------------
def hoy():
    return datetime.date.today().isoformat()


def parse_date(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(str(s)[:10], fmt).date()
        except ValueError:
            continue
    return None


def registrar_historial(db, aid, campo, antes, despues, quien=None):
    """Guarda un cambio en el historial, con quién lo hizo."""
    if str(antes) == str(despues):
        return
    db.execute(
        "INSERT INTO historial (actividad_id,campo,valor_antes,valor_despues,fecha,quien) "
        "VALUES (?,?,?,?,?,?)",
        (aid, campo, "" if antes is None else str(antes),
         "" if despues is None else str(despues),
         datetime.datetime.now().isoformat(timespec="seconds"), quien))


def estatus_por_avance(av):
    if av <= 0:
        return "Pendiente"
    if av >= 100:
        return "Listo"
    return "En proceso"


def dias_restantes():
    fin = parse_date(FECHA_ENTREGA)
    return (fin - datetime.date.today()).days


def evaluar_dependencias_area(db, area, actividad_tipo_o_giro=None):
    """
    Evalúa el estado de las dependencias para un área dada (o para un tipo específico dentro del área).
    Devuelve dict con:
      bloqueada: bool,
      liberada: bool,
      estado: str,
      detalle: str,
      predecesoras_avance: float,
      umbral: int,
      forzada: bool
    """
    if not area:
        return {"bloqueada": False, "liberada": True, "estado": "Sin dependencias", "detalle": "", "forzada": False}
    
    area_norm = sin_acentos(area)
    all_deps = db.execute("SELECT * FROM dependencias").fetchall()
    deps = [d for d in all_deps if sin_acentos(d["area"]) == area_norm or sin_acentos(d["area"]) in area_norm or area_norm in sin_acentos(d["area"])]
    if not deps:
        return {"bloqueada": False, "liberada": True, "estado": "Sin dependencias", "detalle": "", "forzada": False}
    
    # Si se pide para un tipo específico, filtrar dependencias que apliquen a ese tipo sucesor
    if actividad_tipo_o_giro:
        act_tipo_norm = sin_acentos(actividad_tipo_o_giro)
        deps_filtradas = [d for d in deps if sin_acentos(d["tipo_sucesor"]) in act_tipo_norm or act_tipo_norm in sin_acentos(d["tipo_sucesor"])]
        if not deps_filtradas:
            return {"bloqueada": False, "liberada": True, "estado": "Sin dependencias", "detalle": "", "forzada": False}
        deps = deps_filtradas

    # Evaluar las dependencias activas
    for d in deps:
        if d["liberacion_forzada"]:
            return {
                "bloqueada": False,
                "liberada": True,
                "estado": "Liberada (anticipada)",
                "detalle": f"Liberación anticipada por {d['forzada_por'] or 'admin'} ({d['forzada_fecha'] or ''}): {d['forzada_nota'] or ''}",
                "forzada": True,
                "umbral": d["umbral"] or 100,
                "dep_id": d["id"]
            }
        
        # Obtener lista de tipos predecesores
        pred_txt = d["tipos_predecesores"] or ""
        preds = [sin_acentos(p.strip()) for p in pred_txt.split(",") if p.strip()]
        
        # Buscar avance oficial de las actividades predecesoras en esa área
        filas_area = db.execute(
            "SELECT avance, giro, tipo, tipo_partida, partida FROM actividades WHERE area=? AND (aplica IS NULL OR aplica<>'NO')",
            (area,)
        ).fetchall()
        
        avances_pred = []
        for fa in filas_area:
            tps = [sin_acentos(fa["giro"]), sin_acentos(fa["tipo"]), sin_acentos(fa["tipo_partida"]), sin_acentos(fa["partida"])]
            texto_comp = " ".join(tps)
            if any(p in texto_comp for p in preds):
                avances_pred.append(fa["avance"] or 0)
                
        if not avances_pred:
            # Si no hay predecesoras declaradas o encontradas en esa área, continúa evaluando
            continue
            
        prom_avance = sum(avances_pred) / len(avances_pred)
        umbral = d["umbral"] or 100
        
        if prom_avance < umbral:
            return {
                "bloqueada": True,
                "liberada": False,
                "estado": f"Bloqueada (Predecesoras al {prom_avance:.0f}%, requiere {umbral}%)",
                "detalle": f"Esperando que '{pred_txt}' alcance el {umbral}% en {area} (avance actual: {prom_avance:.0f}%)",
                "predecesoras_avance": round(prom_avance, 1),
                "umbral": umbral,
                "forzada": False,
                "dep_id": d["id"]
            }
            
    return {
        "bloqueada": False,
        "liberada": True,
        "estado": "Liberada (cumplió umbral)",
        "detalle": "Predecesoras completadas al umbral requerido",
        "forzada": False
    }


# ----------------------------------------------------------------------------
# Rutas de páginas
# ----------------------------------------------------------------------------
@app.route("/api/tipos_internos", methods=["GET"])
@requiere_login
def api_tipos_internos():
    db = get_db()
    rows = db.execute("SELECT nombre FROM tipos_internos ORDER BY id").fetchall()
    return jsonify([r["nombre"] for r in rows])


@app.route("/api/tipos_internos", methods=["POST"])
@requiere_gestor
def api_tipos_internos_add():
    db = get_db()
    nombre = (request.get_json().get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "Escribe un nombre"}), 400
    db.execute("INSERT OR IGNORE INTO tipos_internos (nombre) VALUES (?)", (nombre,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/involucrados")
def involucrados_page():
    return render_template("involucrados.html")


@app.route("/api/involucrados/catalogo", methods=["GET"])
@requiere_gestor
def api_inv_catalogo():
    db = get_db()
    rows = db.execute("SELECT nombre FROM involucrados ORDER BY nombre").fetchall()
    return jsonify([r["nombre"] for r in rows])


@app.route("/api/involucrados/catalogo", methods=["POST"])
@requiere_gestor
def api_inv_catalogo_add():
    db = get_db()
    nombre = (request.get_json().get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "Escribe un nombre"}), 400
    db.execute("INSERT OR IGNORE INTO involucrados (nombre) VALUES (?)", (nombre,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/involucrados/por_area", methods=["GET"])
@requiere_gestor
def api_inv_por_area():
    """Devuelve, por cada área, qué involucrados tiene encendidos (y el catálogo completo)."""
    db = get_db()
    catalogo = [r["nombre"] for r in db.execute("SELECT nombre FROM involucrados ORDER BY nombre").fetchall()]
    areas = db.execute(
        "SELECT DISTINCT bloque, area FROM actividades WHERE area IS NOT NULL AND area<>'' "
        "ORDER BY bloque, area").fetchall()
    rel = {}
    for r in db.execute("SELECT area, involucrado FROM area_involucrados").fetchall():
        rel.setdefault(r["area"], set()).add(r["involucrado"])
    salida = []
    for a in areas:
        encendidos = sorted(rel.get(a["area"], set()))
        salida.append({"bloque": a["bloque"] or "", "area": a["area"], "encendidos": encendidos})
    return jsonify({"catalogo": catalogo, "areas": salida})


@app.route("/api/involucrados/toggle", methods=["POST"])
@requiere_gestor
def api_inv_toggle():
    """Prende o apaga un involucrado en una área."""
    db = get_db()
    data = request.get_json()
    area = data.get("area"); who = data.get("involucrado")
    if not area or not who:
        return jsonify({"error": "faltan datos"}), 400
    existe = db.execute("SELECT 1 FROM area_involucrados WHERE area=? AND involucrado=?", (area, who)).fetchone()
    if existe:
        db.execute("DELETE FROM area_involucrados WHERE area=? AND involucrado=?", (area, who))
        estado = "off"
    else:
        db.execute("INSERT OR IGNORE INTO area_involucrados (area,involucrado) VALUES (?,?)", (area, who))
        estado = "on"
    db.commit()
    return jsonify({"ok": True, "estado": estado})


@app.route("/login")
def login_page():
    return render_template("login.html")


@app.route("/")
def index():
    # si no hay sesión, al login; si es proveedor, a su portal
    if "usuario" not in session:
        return redirect("/login")
    if not es_gestor(session.get("rol")):
        return redirect("/portal")
    return render_template("index.html")


@app.route("/reporte")
def reporte():
    return render_template("reporte.html")


@app.route("/reportes")
def reportes():
    return render_template("reportes.html")


@app.route("/expediente")
def expediente_page():
    return render_template("expediente.html")


# ----------------------------------------------------------------------------
# API — Actividades
# ----------------------------------------------------------------------------
@app.route("/api/actividades")
def api_actividades():
    db = get_db()
    q = "SELECT * FROM actividades"
    cond = []
    args = []
    bloque = request.args.get("bloque")
    area = request.args.get("area")
    giro = request.args.get("giro")
    proveedor = request.args.get("proveedor")
    tipo_partida = request.args.get("tipo_partida")
    estatus = request.args.get("estatus")
    buscar = request.args.get("buscar")
    # mundo: 'obra' (default), 'interno' o 'todos'
    mundo = request.args.get("mundo", "obra")
    if mundo != "todos":
        cond.append("(mundo = ? OR (mundo IS NULL AND ? = 'obra'))")
        args += [mundo, mundo]
    if bloque:
        cond.append("bloque = ?"); args.append(bloque)
    if area:
        cond.append("area = ?"); args.append(area)
    if giro:
        cond.append("giro = ?"); args.append(giro)
    if proveedor:
        cond.append("proveedor = ?"); args.append(proveedor)
    if tipo_partida:
        cond.append("tipo_partida = ?"); args.append(tipo_partida)
    if estatus:
        cond.append("estatus = ?"); args.append(estatus)
    if buscar:
        b = "%" + sin_acentos(buscar) + "%"
        cond.append("(sinac(partida) LIKE ? OR sinac(area) LIKE ? OR sinac(proveedor) LIKE ? OR sinac(bloque) LIKE ? OR sinac(codigo) LIKE ?)")
        args += [b] * 5
    # por defecto no mostramos las rechazadas en la tabla principal
    if request.args.get("incluir_rechazadas") != "1":
        cond.append("(estado_val IS NULL OR estado_val<>'rechazada')")
    if cond:
        q += " WHERE " + " AND ".join(cond)
    q += " ORDER BY id"
    rows = db.execute(q, args).fetchall()
    salida = []
    for r in rows:
        d = dict(r)
        dep_st = evaluar_dependencias_area(db, r["area"], r["tipo_partida"] or r["giro"])
        d["dep_bloqueada"] = dep_st["bloqueada"]
        d["dep_estado"] = dep_st["estado"]
        d["dep_detalle"] = dep_st["detalle"]
        salida.append(d)
    return jsonify(salida)


@app.route("/api/actividad/<int:aid>", methods=["GET"])
def api_actividad(aid):
    db = get_db()
    r = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    if not r:
        return jsonify({"error": "no existe"}), 404
    return jsonify(dict(r))


@app.route("/api/actividad/<int:aid>/historial")
def api_actividad_historial(aid):
    db = get_db()
    a = db.execute("SELECT codigo,partida,area,bloque FROM actividades WHERE id=?", (aid,)).fetchone()
    rows = db.execute(
        "SELECT campo,valor_antes,valor_despues,fecha,quien FROM historial "
        "WHERE actividad_id=? ORDER BY fecha DESC, id DESC", (aid,)).fetchall()
    return jsonify({
        "actividad": dict(a) if a else None,
        "historial": [dict(r) for r in rows],
    })


@app.route("/api/actividad/<int:aid>", methods=["PUT"])
@requiere_gestor
def api_actualizar(aid):
    db = get_db()
    data = request.get_json()
    actual = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    if not actual:
        return jsonify({"error": "no existe"}), 404
    campos = ["proveedor", "partida", "giro", "aplica", "avance", "f_inicio",
              "f_fin", "duracion_dias", "estatus", "depende_de", "notas", "tipo",
              "area", "bloque", "tipo_partida", "definido",
              "causa_retraso", "nota_proveedor",
              "mundo", "departamento", "tipo_interno"]
    # si mandan causa_retraso, sellamos quién y cuándo
    if "causa_retraso" in data and (data.get("causa_retraso") or "").strip():
        data["causa_por"] = data.get("causa_por") or "Registrado en plataforma"
        data["causa_fecha"] = datetime.date.today().isoformat()
        campos += ["causa_por", "causa_fecha"]
    updates = []
    args = []
    for c in campos:
        if c in data:
            antes = actual[c]
            despues = data[c]
            if str(antes) != str(despues):
                db.execute(
                    "INSERT INTO historial (actividad_id,campo,valor_antes,valor_despues,fecha) VALUES (?,?,?,?,?)",
                    (aid, c, str(antes), str(despues), datetime.datetime.now().isoformat(timespec="seconds")),
                )
            updates.append(f"{c}=?")
            args.append(despues)
    # auto estatus si mandan avance y no mandan estatus explícito
    if "avance" in data and "estatus" not in data:
        updates.append("estatus=?")
        args.append(estatus_por_avance(int(data["avance"] or 0)))
    updates.append("actualizado=?")
    args.append(datetime.datetime.now().isoformat(timespec="seconds"))
    args.append(aid)
    db.execute(f"UPDATE actividades SET {','.join(updates)} WHERE id=?", args)
    db.commit()
    r = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    return jsonify(dict(r))


@app.route("/api/actividad", methods=["POST"])
@requiere_gestor
def api_crear():
    db = get_db()
    data = request.get_json()
    cur = db.execute(
        """INSERT INTO actividades
        (codigo,bloque,area,giro,proveedor,partida,tipo,tipo_partida,aplica,avance,
         f_inicio,f_fin,duracion_dias,estatus,depende_de,definido,causa_retraso,notas,
         mundo,departamento,tipo_interno,actualizado)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            data.get("codigo"), data.get("bloque"), data.get("area"),
            data.get("giro"), data.get("proveedor"), data.get("partida"),
            data.get("tipo"), data.get("tipo_partida", "Construcción"),
            data.get("aplica", "SÍ"),
            int(data.get("avance", 0) or 0), data.get("f_inicio"),
            data.get("f_fin"), data.get("duracion_dias"),
            data.get("estatus", "Pendiente"), data.get("depende_de"),
            data.get("definido", "NO"), data.get("causa_retraso"), data.get("notas"),
            data.get("mundo", "obra"), data.get("departamento"), data.get("tipo_interno"),
            datetime.datetime.now().isoformat(timespec="seconds"),
        ),
    )
    db.commit()
    r = db.execute("SELECT * FROM actividades WHERE id=?", (cur.lastrowid,)).fetchone()
    return jsonify(dict(r))


@app.route("/api/actividad/<int:aid>", methods=["DELETE"])
@requiere_admin
def api_borrar(aid):
    db = get_db()
    db.execute("DELETE FROM actividades WHERE id=?", (aid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/actividades/borrar", methods=["POST"])
@requiere_admin
def api_borrar_multiple():
    db = get_db()
    data = request.get_json()
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "sin ids"}), 400
    marcas = ",".join("?" * len(ids))
    db.execute(f"DELETE FROM actividades WHERE id IN ({marcas})", ids)
    db.commit()
    return jsonify({"ok": True, "borradas": len(ids)})


# ----------------------------------------------------------------------------
# API — Catálogos y resumen
# ----------------------------------------------------------------------------
@app.route("/api/catalogos")
def api_catalogos():
    db = get_db()
    def distintos(col):
        rows = db.execute(
            f"SELECT DISTINCT {col} FROM actividades WHERE {col} IS NOT NULL AND {col}<>'' ORDER BY {col}"
        ).fetchall()
        return [r[0] for r in rows]
    # departamentos = los del catálogo de involucrados + los que ya tienen actividades internas
    invol = [r[0] for r in db.execute("SELECT nombre FROM involucrados ORDER BY nombre").fetchall()]
    dept_act = distintos("departamento")
    departamentos = sorted(set(invol) | set(dept_act))
    # mapa bloque -> áreas (para el candado estricto en el formulario)
    mapa = {}
    for r in db.execute(
        "SELECT DISTINCT bloque, area FROM actividades "
        "WHERE bloque IS NOT NULL AND bloque<>'' AND area IS NOT NULL AND area<>'' "
        "ORDER BY bloque, area").fetchall():
        mapa.setdefault(r["bloque"], []).append(r["area"])
    return jsonify({
        "bloques": distintos("bloque"),
        "areas": distintos("area"),
        "giros": distintos("giro"),
        "proveedores": distintos("proveedor"),
        "departamentos": departamentos,
        "mapa_bloque_areas": mapa,
        "tipos_partida": ["Construcción", "Mobiliario y equipo", "Puesta en marcha", "Detalles finales"],
        "estatus": ["Pendiente", "En proceso", "Listo", "Post-apertura"],
    })


# ----------------------------------------------------------------------------
# API — Catálogo de proveedores (con tipo interno/externo y función)
# ----------------------------------------------------------------------------
@app.route("/api/proveedores", methods=["GET"])
def api_proveedores():
    db = get_db()
    rows = db.execute("SELECT * FROM proveedores ORDER BY nombre").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/proveedores", methods=["POST"])
def api_crear_proveedor():
    db = get_db()
    data = request.get_json()
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "nombre requerido"}), 400
    try:
        db.execute(
            "INSERT INTO proveedores (nombre,tipo,funcion) VALUES (?,?,?)",
            (nombre, data.get("tipo", "Externo"), data.get("funcion", "")),
        )
        db.commit()
    except sqlite3.IntegrityError:
        # ya existe: actualiza tipo y función
        db.execute(
            "UPDATE proveedores SET tipo=?, funcion=? WHERE nombre=?",
            (data.get("tipo", "Externo"), data.get("funcion", ""), nombre),
        )
        db.commit()
    r = db.execute("SELECT * FROM proveedores WHERE nombre=?", (nombre,)).fetchone()
    return jsonify(dict(r))


# ----------------------------------------------------------------------------
# API — Catálogo genérico (bloque / area / giro) con nota
# ----------------------------------------------------------------------------
@app.route("/api/catalogo/<clase>", methods=["POST"])
def api_crear_catalogo(clase):
    if clase not in ("bloque", "area", "giro"):
        return jsonify({"error": "clase inválida"}), 400
    db = get_db()
    data = request.get_json()
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "nombre requerido"}), 400
    try:
        db.execute(
            "INSERT INTO catalogo (clase,nombre,nota) VALUES (?,?,?)",
            (clase, nombre, data.get("nota", "")),
        )
        db.commit()
    except sqlite3.IntegrityError:
        db.execute(
            "UPDATE catalogo SET nota=? WHERE clase=? AND nombre=?",
            (data.get("nota", ""), clase, nombre),
        )
        db.commit()
    return jsonify({"ok": True, "clase": clase, "nombre": nombre})


# ----------------------------------------------------------------------------
# API — Causas de retraso (catálogo: 7 base + agregadas por el usuario)
# ----------------------------------------------------------------------------
@app.route("/api/causas", methods=["GET"])
def api_causas():
    db = get_db()
    mundo = request.args.get("mundo", "obra")
    rows = db.execute(
        "SELECT nombre FROM causas WHERE (mundo = ? OR (mundo IS NULL AND ? = 'obra')) "
        "ORDER BY base DESC, nombre", (mundo, mundo)).fetchall()
    return jsonify([r[0] for r in rows])


@app.route("/api/causas", methods=["POST"])
def api_crear_causa():
    db = get_db()
    data = request.get_json()
    nombre = (data.get("nombre") or "").strip()
    mundo = data.get("mundo", "obra")
    if not nombre:
        return jsonify({"error": "nombre requerido"}), 400
    db.execute("INSERT OR IGNORE INTO causas (nombre,base,mundo) VALUES (?,0,?)", (nombre, mundo))
    db.commit()
    rows = db.execute(
        "SELECT nombre FROM causas WHERE (mundo = ? OR (mundo IS NULL AND ? = 'obra')) "
        "ORDER BY base DESC, nombre", (mundo, mundo)).fetchall()
    return jsonify([r[0] for r in rows])


# ----------------------------------------------------------------------------
# API — Resumen de causas AGREGADAS (para el reporte a dirección, sin señalar)
# ----------------------------------------------------------------------------
@app.route("/api/resumen_causas")
def api_resumen_causas():
    db = get_db()
    rows = db.execute(
        """SELECT causa_retraso, COUNT(*) n
           FROM actividades
           WHERE causa_retraso IS NOT NULL AND causa_retraso<>''
             AND (aplica IS NULL OR aplica<>'NO')
           GROUP BY causa_retraso ORDER BY n DESC"""
    ).fetchall()
    total = sum(r["n"] for r in rows)
    salida = [{"causa": r["causa_retraso"], "partidas": r["n"],
               "porcentaje": round(100 * r["n"] / total, 1) if total else 0}
              for r in rows]
    return jsonify({"total_con_causa": total, "causas": salida})


# ----------------------------------------------------------------------------
# API — Expediente de proveedores (datos de empresa y contacto)
# ----------------------------------------------------------------------------
@app.route("/api/expediente")
def api_expediente():
    """Lista proveedores usados en actividades + su ficha (si existe)."""
    db = get_db()
    # proveedores que aparecen en actividades
    usados = db.execute(
        "SELECT proveedor, COUNT(*) n, ROUND(AVG(avance),1) av FROM actividades "
        "WHERE proveedor IS NOT NULL AND proveedor<>'' GROUP BY proveedor ORDER BY proveedor"
    ).fetchall()
    fichas = {r["nombre"]: dict(r) for r in db.execute("SELECT * FROM proveedores").fetchall()}
    salida = []
    for u in usados:
        f = fichas.get(u["proveedor"], {})
        salida.append({
            "nombre": u["proveedor"], "partidas": u["n"], "avance": u["av"],
            "empresa": f.get("empresa", ""), "tipo": f.get("tipo", ""),
            "funcion": f.get("funcion", ""), "contacto": f.get("contacto", ""),
            "telefono": f.get("telefono", ""), "correo": f.get("correo", ""),
            "notas": f.get("notas", ""),
        })
    return jsonify(salida)


@app.route("/api/expediente", methods=["POST"])
def api_guardar_expediente():
    db = get_db()
    data = request.get_json()
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "nombre requerido"}), 400
    existe = db.execute("SELECT id FROM proveedores WHERE nombre=?", (nombre,)).fetchone()
    campos = ("empresa", "tipo", "funcion", "contacto", "telefono", "correo", "notas")
    if existe:
        sets = ",".join(f"{c}=?" for c in campos)
        db.execute(f"UPDATE proveedores SET {sets} WHERE nombre=?",
                   [data.get(c, "") for c in campos] + [nombre])
    else:
        db.execute(
            "INSERT INTO proveedores (nombre,empresa,tipo,funcion,contacto,telefono,correo,notas) "
            "VALUES (?,?,?,?,?,?,?,?)",
            [nombre] + [data.get(c, "") for c in campos])
    db.commit()
    return jsonify({"ok": True})


# ----------------------------------------------------------------------------
# API — Snapshots semanales (comparar semana vs semana)
# ----------------------------------------------------------------------------
@app.route("/api/snapshot", methods=["POST"])
def api_snapshot():
    """Guarda una foto del avance de todas las partidas hoy."""
    db = get_db()
    hoy_d = datetime.date.today()
    semana = hoy_d.strftime("%G-W%V")  # año-semana ISO
    fecha = hoy_d.isoformat()
    # borra foto previa de la misma semana para no duplicar
    db.execute("DELETE FROM snapshots WHERE semana=?", (semana,))
    rows = db.execute("SELECT id, avance FROM actividades").fetchall()
    for r in rows:
        db.execute(
            "INSERT INTO snapshots (semana,fecha,actividad_id,avance) VALUES (?,?,?,?)",
            (semana, fecha, r["id"], r["avance"] or 0))
    db.commit()
    return jsonify({"ok": True, "semana": semana, "partidas": len(rows)})


@app.route("/api/comparar_semanas")
def api_comparar_semanas():
    """Compara el avance actual contra la última foto guardada."""
    db = get_db()
    semanas = db.execute(
        "SELECT DISTINCT semana FROM snapshots ORDER BY semana DESC").fetchall()
    if not semanas:
        return jsonify({"hay_foto": False})
    ultima = semanas[0]["semana"]
    # avance guardado en la última foto, por proveedor
    prev = db.execute(
        """SELECT a.proveedor prov, AVG(s.avance) av
           FROM snapshots s JOIN actividades a ON a.id=s.actividad_id
           WHERE s.semana=? GROUP BY a.proveedor""", (ultima,)).fetchall()
    prev_map = {r["prov"]: r["av"] for r in prev}
    # avance actual por proveedor
    act = db.execute(
        "SELECT proveedor prov, AVG(avance) av FROM actividades GROUP BY proveedor").fetchall()
    salida = []
    for r in act:
        p = r["prov"] or "— Sin proveedor —"
        antes = prev_map.get(r["prov"], 0) or 0
        ahora = r["av"] or 0
        salida.append({
            "proveedor": p, "antes": round(antes, 1), "ahora": round(ahora, 1),
            "cambio": round(ahora - antes, 1),
        })
    salida.sort(key=lambda x: x["cambio"])
    return jsonify({"hay_foto": True, "semana_previa": ultima, "proveedores": salida})



# ============================================================
#  HISTÓRICO — consultar cómo estaba la obra en un día pasado
#  v1.5 · Reconstruye el pasado desde las fotos semanales y,
#  si no hay foto, deshaciendo los cambios del historial.
# ============================================================

def _fotos_disponibles(db):
    """Fechas en las que hay foto guardada, de la más nueva a la más vieja."""
    return [dict(r) for r in db.execute(
        "SELECT semana, MIN(fecha) fecha, COUNT(*) partidas "
        "FROM snapshots GROUP BY semana ORDER BY fecha DESC")]


def _avance_a_fecha(db, fecha, mundo="obra"):
    """Devuelve {actividad_id: avance} tal como estaba ESE DÍA.
    1) Si hay foto de ese día o anterior, la usa.
    2) Si no, reconstruye: parte del avance de hoy y deshace
       los cambios que ocurrieron DESPUÉS de la fecha pedida."""
    fila = db.execute(
        "SELECT semana, fecha FROM snapshots WHERE fecha<=? "
        "ORDER BY fecha DESC LIMIT 1", (fecha,)).fetchone()
    origen = "reconstruido"
    mapa = {}
    if fila:
        origen = "foto"
        for r in db.execute("SELECT actividad_id, avance FROM snapshots WHERE semana=?",
                            (fila["semana"],)):
            mapa[r["actividad_id"]] = r["avance"] or 0
    # actividades vigentes del mundo pedido
    acts = db.execute(
        "SELECT id, avance FROM actividades "
        "WHERE (aplica IS NULL OR aplica<>'NO') "
        "AND (estado_val IS NULL OR estado_val='validado') "
        "AND (mundo = ? OR (mundo IS NULL AND ? = 'obra'))",
        (mundo, mundo)).fetchall()
    for a in acts:
        if a["id"] in mapa:
            continue
        # reconstruir desde el historial: el primer cambio POSTERIOR a la
        # fecha nos dice cuánto valía ANTES de ese cambio
        h = db.execute(
            "SELECT valor_antes FROM historial "
            "WHERE actividad_id=? AND (campo LIKE '%avance%') AND fecha> ? "
            "ORDER BY fecha ASC LIMIT 1", (a["id"], fecha)).fetchone()
        if h and h["valor_antes"] not in (None, "", "None"):
            try:
                mapa[a["id"]] = int(float(h["valor_antes"]))
            except (ValueError, TypeError):
                mapa[a["id"]] = a["avance"] or 0
        else:
            # no hubo cambios después de esa fecha: sigue igual que hoy
            mapa[a["id"]] = a["avance"] or 0
    return mapa, origen, (fila["fecha"] if fila else None)


@app.route("/api/historico/fechas")
@requiere_gestor
def api_historico_fechas():
    db = get_db()
    fotos = _fotos_disponibles(db)
    rango = db.execute(
        "SELECT MIN(substr(fecha,1,10)) ini, MAX(substr(fecha,1,10)) fin FROM historial").fetchone()
    return jsonify({
        "fotos": fotos,
        "historial_desde": rango["ini"] if rango else None,
        "historial_hasta": rango["fin"] if rango else None,
        "hoy": datetime.date.today().isoformat(),
    })


@app.route("/api/historico")
@requiere_gestor
def api_historico():
    """Cómo estaba la obra en una fecha dada."""
    fecha = request.args.get("fecha") or datetime.date.today().isoformat()
    mundo = request.args.get("mundo", "obra")
    db = get_db()
    mapa, origen, fecha_foto = _avance_a_fecha(db, fecha, mundo)
    filas = db.execute(
        "SELECT id, codigo, bloque, area, giro, proveedor, departamento, mundo, partida, avance "
        "FROM actividades WHERE (aplica IS NULL OR aplica<>'NO') "
        "AND (estado_val IS NULL OR estado_val='validado') "
        "AND (mundo = ? OR (mundo IS NULL AND ? = 'obra'))",
        (mundo, mundo)).fetchall()
    if not filas:
        return jsonify({"total": 0, "fecha": fecha, "origen": origen})
    por_bloque, por_resp, detalle = {}, {}, []
    suma = 0
    for f in filas:
        ent = mapa.get(f["id"], 0)
        hoy_av = f["avance"] or 0
        suma += ent
        b = f["bloque"] or "— Sin bloque —"
        quien = (f["departamento"] if f["mundo"] == "interno" else f["proveedor"]) or "— Sin asignar —"
        por_bloque.setdefault(b, []).append(ent)
        por_resp.setdefault(quien, []).append((ent, hoy_av))
        detalle.append({
            "codigo": f["codigo"], "bloque": b, "area": f["area"],
            "responsable": quien, "partida": f["partida"],
            "avance_fecha": ent, "avance_hoy": hoy_av, "cambio": hoy_av - ent,
        })
    n = len(filas)
    bloques = [{"bloque": k, "total": len(v), "avance": round(sum(v) / len(v), 1)}
               for k, v in por_bloque.items()]
    bloques.sort(key=lambda x: -x["avance"])
    resp = []
    for k, v in por_resp.items():
        ent = sum(x[0] for x in v) / len(v)
        hoy_ = sum(x[1] for x in v) / len(v)
        resp.append({"responsable": k, "total": len(v), "avance": round(ent, 1),
                     "avance_hoy": round(hoy_, 1), "cambio": round(hoy_ - ent, 1)})
    resp.sort(key=lambda x: -x["total"])
    detalle.sort(key=lambda x: -abs(x["cambio"]))
    return jsonify({
        "fecha": fecha, "origen": origen, "fecha_foto": fecha_foto,
        "total": n, "avance_global": round(suma / n, 1),
        "avance_hoy": round(sum((f["avance"] or 0) for f in filas) / n, 1),
        "bloques": bloques, "responsables": resp, "detalle": detalle[:400],
    })


@app.route("/api/historico/comparar")
@requiere_gestor
def api_historico_comparar():
    """Cuánto se movió cada responsable entre dos fechas."""
    desde = request.args.get("desde")
    hasta = request.args.get("hasta") or datetime.date.today().isoformat()
    mundo = request.args.get("mundo", "obra")
    if not desde:
        return jsonify({"error": "Falta la fecha inicial"}), 400
    db = get_db()
    m1, o1, _ = _avance_a_fecha(db, desde, mundo)
    m2, o2, _ = _avance_a_fecha(db, hasta, mundo)
    filas = db.execute(
        "SELECT id, bloque, proveedor, departamento, mundo FROM actividades "
        "WHERE (aplica IS NULL OR aplica<>'NO') "
        "AND (estado_val IS NULL OR estado_val='validado') "
        "AND (mundo = ? OR (mundo IS NULL AND ? = 'obra'))",
        (mundo, mundo)).fetchall()
    if not filas:
        return jsonify({"total": 0})
    por_resp, por_bloque = {}, {}
    for f in filas:
        quien = (f["departamento"] if f["mundo"] == "interno" else f["proveedor"]) or "— Sin asignar —"
        a1, a2 = m1.get(f["id"], 0), m2.get(f["id"], 0)
        por_resp.setdefault(quien, []).append((a1, a2))
        por_bloque.setdefault(f["bloque"] or "— Sin bloque —", []).append((a1, a2))
    def arma(d, etiqueta):
        out = []
        for k, v in d.items():
            p1 = sum(x[0] for x in v) / len(v)
            p2 = sum(x[1] for x in v) / len(v)
            out.append({etiqueta: k, "total": len(v), "antes": round(p1, 1),
                        "despues": round(p2, 1), "cambio": round(p2 - p1, 1)})
        out.sort(key=lambda x: x["cambio"])
        return out
    g1 = sum(m1.get(f["id"], 0) for f in filas) / len(filas)
    g2 = sum(m2.get(f["id"], 0) for f in filas) / len(filas)
    return jsonify({
        "desde": desde, "hasta": hasta, "origen_desde": o1, "origen_hasta": o2,
        "global_antes": round(g1, 1), "global_despues": round(g2, 1),
        "global_cambio": round(g2 - g1, 1),
        "responsables": arma(por_resp, "responsable"),
        "bloques": arma(por_bloque, "bloque"),
    })


def _foto_automatica():
    """Guarda la foto de la semana si todavía no existe. Se llama al arrancar,
    para que el histórico no dependa de que alguien se acuerde de picarle."""
    try:
        con = sqlite3.connect(DB_PATH)
        con.row_factory = sqlite3.Row
        hoy_d = datetime.date.today()
        semana = hoy_d.strftime("%G-W%V")
        ya = con.execute("SELECT COUNT(*) FROM snapshots WHERE semana=?", (semana,)).fetchone()[0]
        if ya == 0:
            filas = con.execute("SELECT id, avance FROM actividades").fetchall()
            for r in filas:
                con.execute(
                    "INSERT INTO snapshots (semana,fecha,actividad_id,avance) VALUES (?,?,?,?)",
                    (semana, hoy_d.isoformat(), r["id"], r["avance"] or 0))
            con.commit()
            print(f"  Foto semanal automatica guardada ({semana}, {len(filas)} partidas)")
        con.close()
    except Exception as e:
        print("  (no se pudo guardar la foto automatica:", e, ")")


@app.route("/historico")
def historico_page():
    return render_template("historico.html")


def _resumen_data():
    db = get_db()
    mundo = request.args.get("mundo", "obra")
    rows = db.execute(
        "SELECT * FROM actividades WHERE (aplica IS NULL OR aplica<>'NO') "
        "AND (estado_val IS NULL OR estado_val='validado') "
        "AND (mundo = ? OR (mundo IS NULL AND ? = 'obra'))",
        (mundo, mundo)
    ).fetchall()
    total = len(rows)
    if total == 0:
        return {"total": 0}
    suma_av = sum((r["avance"] or 0) for r in rows)
    avance_global = round(suma_av / total, 1)

    # por estatus
    est = {}
    for r in rows:
        e = r["estatus"] or "Pendiente"
        est[e] = est.get(e, 0) + 1

    # por bloque
    bloques = {}
    for r in rows:
        b = r["bloque"] or "—"
        bloques.setdefault(b, {"total": 0, "suma": 0})
        bloques[b]["total"] += 1
        bloques[b]["suma"] += (r["avance"] or 0)
    bloques_out = []
    for b, d in bloques.items():
        bloques_out.append({
            "bloque": b, "total": d["total"],
            "avance": round(d["suma"] / d["total"], 1),
        })
    bloques_out.sort(key=lambda x: x["avance"])

    # por giro
    giros = {}
    for r in rows:
        gg = r["giro"] or "—"
        giros.setdefault(gg, {"total": 0, "suma": 0})
        giros[gg]["total"] += 1
        giros[gg]["suma"] += (r["avance"] or 0)
    giros_out = [{"giro": k, "total": v["total"], "avance": round(v["suma"]/v["total"],1)} for k, v in giros.items()]
    giros_out.sort(key=lambda x: x["avance"])

    # por tipo de partida (Construcción, Mobiliario y equipo, etc.)
    tipos = {}
    for r in rows:
        tp = (r["tipo_partida"] if "tipo_partida" in r.keys() else None) or "Construcción"
        tipos.setdefault(tp, {"total": 0, "suma": 0})
        tipos[tp]["total"] += 1
        tipos[tp]["suma"] += (r["avance"] or 0)
    tipos_out = [{"tipo": k, "total": v["total"], "avance": round(v["suma"]/v["total"],1)} for k, v in tipos.items()]
    tipos_out.sort(key=lambda x: x["avance"])

    # retrasos: fecha fin pasada y avance < 100
    hoy_d = datetime.date.today()
    retrasadas = []
    en_riesgo = []
    for r in rows:
        fin = parse_date(r["f_fin"])
        av = r["avance"] or 0
        if fin and av < 100:
            dif = (fin - hoy_d).days
            item = {
                "id": r["id"], "area": r["area"], "partida": r["partida"],
                "proveedor": r["proveedor"], "giro": r["giro"],
                "f_fin": r["f_fin"], "avance": av, "dias": dif,
            }
            if dif < 0:
                retrasadas.append(item)
            elif dif <= 7:
                en_riesgo.append(item)
    retrasadas.sort(key=lambda x: x["dias"])
    en_riesgo.sort(key=lambda x: x["dias"])

    return {
        "total": total,
        "avance_global": avance_global,
        "estatus": est,
        "bloques": bloques_out,
        "giros": giros_out,
        "tipos": tipos_out,
        "retrasadas": retrasadas,
        "en_riesgo": en_riesgo,
        "dias_restantes": dias_restantes(),
        "fecha_entrega": FECHA_ENTREGA,
        "fecha_hoy": hoy_d.isoformat(),
    }


@app.route("/api/resumen")
def api_resumen():
    return jsonify(_resumen_data())


# ----------------------------------------------------------------------------
# API — Resumen por proveedor (para el centro de reportes)
# ----------------------------------------------------------------------------
@app.route("/api/resumen_proveedores")
def api_resumen_proveedores():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM actividades WHERE aplica IS NULL OR aplica<>'NO'"
    ).fetchall()
    hoy_d = datetime.date.today()
    prov = {}
    for r in rows:
        p = r["proveedor"] or "— Sin proveedor —"
        prov.setdefault(p, {"total": 0, "suma": 0, "listas": 0, "retrasadas": 0})
        prov[p]["total"] += 1
        prov[p]["suma"] += (r["avance"] or 0)
        if (r["avance"] or 0) >= 100:
            prov[p]["listas"] += 1
        fin = parse_date(r["f_fin"])
        if fin and (r["avance"] or 0) < 100 and (fin - hoy_d).days < 0:
            prov[p]["retrasadas"] += 1
    salida = []
    for p, d in prov.items():
        salida.append({
            "proveedor": p, "total": d["total"],
            "avance": round(d["suma"] / d["total"], 1) if d["total"] else 0,
            "listas": d["listas"], "retrasadas": d["retrasadas"],
        })
    salida.sort(key=lambda x: x["avance"])
    return jsonify(salida)


def _filtrar_actividades(args):
    """Devuelve filas filtradas según los mismos parámetros de la tabla."""
    db = get_db()
    q = "SELECT * FROM actividades"
    cond, a = [], []
    for campo in ("bloque", "area", "giro", "proveedor", "tipo_partida", "estatus"):
        v = args.get(campo)
        if v:
            cond.append(f"{campo} = ?"); a.append(v)
    if args.get("solo_retraso") == "1":
        pass  # se filtra abajo con fecha
    if cond:
        q += " WHERE " + " AND ".join(cond)
    q += " ORDER BY area, id"
    rows = db.execute(q, a).fetchall()
    if args.get("solo_retraso") == "1":
        hoy_d = datetime.date.today()
        rows = [r for r in rows if parse_date(r["f_fin"]) and (r["avance"] or 0) < 100
                and (parse_date(r["f_fin"]) - hoy_d).days < 0]
    return rows


# ----------------------------------------------------------------------------
# Hoja para proveedor — Excel (para llenar) y PDF (para anotar a mano)
# ----------------------------------------------------------------------------
@app.route("/api/hoja_proveedor.xlsx")
def hoja_proveedor_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    prov = request.args.get("proveedor", "")
    titulo_txt = request.args.get("titulo", "")
    rows = _filtrar_actividades(request.args)
    hoy_d = datetime.date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actualización"
    encabezado = titulo_txt or (f"HOJA DE ACTUALIZACIÓN DE AVANCE — {prov}" if prov else "REPORTE DE ACTIVIDADES")
    ws["A1"] = encabezado
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Fecha: {hoy_d.strftime('%d/%m/%Y')}   ·   Favor de anotar avance, fecha compromiso y observaciones."
    ws["A2"].font = Font(size=9, italic=True, color="555555")
    ws.merge_cells("A2:H2")
    cols = ["Código", "Área", "Partida", "Avance actual",
            "NUEVO avance %", "Fecha compromiso", "¿Retrasada?", "Observaciones"]
    hdr = 4
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=hdr, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    rojo = PatternFill("solid", fgColor="F8D7DA")
    i = hdr + 1
    for r in rows:
        fin = parse_date(r["f_fin"])
        retrasada = fin and (r["avance"] or 0) < 100 and (fin - hoy_d).days < 0
        vals = [r["codigo"], r["area"], r["partida"], f"{r['avance'] or 0}%",
                "", r["f_fin"] or "", "SÍ" if retrasada else "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bd
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if retrasada:
                cell.fill = rojo
        i += 1
    anchos = [10, 22, 40, 12, 13, 15, 11, 30]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"
    safe = "".join(ch for ch in (prov or "proveedor") if ch.isalnum() or ch in " _-")[:30].strip() or "proveedor"
    out = os.path.join(BASE_DIR, "data", "hoja_proveedor.xlsx")
    wb.save(out)
    return send_file(out, as_attachment=True, download_name=f"HOJA_{safe}_{hoy()}.xlsx")


@app.route("/api/hoja_proveedor.pdf")
def hoja_proveedor_pdf():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    prov = request.args.get("proveedor", "")
    titulo_txt = request.args.get("titulo", "")
    rows = _filtrar_actividades(request.args)
    hoy_d = datetime.date.today()
    out = os.path.join(BASE_DIR, "data", "hoja_proveedor.pdf")
    doc = SimpleDocTemplate(out, pagesize=landscape(letter),
                            leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle("t", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#1F4E78"))
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=7, leading=8)
    encabezado = titulo_txt or (f"Hoja de actualización de avance — {prov}" if prov else "Reporte de actividades")
    elems = [Paragraph(encabezado, titulo),
             Paragraph(f"Fecha: {hoy_d.strftime('%d/%m/%Y')} · Anote avance, fecha compromiso y observaciones.", styles["Normal"]),
             Spacer(1, 8)]
    data = [["Código", "Área", "Partida", "Avance\nactual", "NUEVO\navance", "Fecha\ncompromiso", "Retra-\nsada", "Observaciones"]]
    filas_retraso = []
    for idx, r in enumerate(rows):
        fin = parse_date(r["f_fin"])
        retrasada = fin and (r["avance"] or 0) < 100 and (fin - hoy_d).days < 0
        data.append([r["codigo"] or "", Paragraph(r["area"] or "", small),
                     Paragraph(r["partida"] or "", small), f"{r['avance'] or 0}%",
                     "______", r["f_fin"] or "__________", "SÍ" if retrasada else "",
                     "________________"])
        if retrasada:
            filas_retraso.append(idx + 1)
    t = Table(data, colWidths=[2*cm, 3.5*cm, 6*cm, 1.5*cm, 1.6*cm, 2.2*cm, 1.3*cm, 5*cm], repeatRows=1)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7F8")]),
    ]
    for fr in filas_retraso:
        estilo.append(("BACKGROUND", (0, fr), (-1, fr), colors.HexColor("#F8D7DA")))
    t.setStyle(TableStyle(estilo))
    elems.append(t)
    doc.build(elems)
    safe = "".join(ch for ch in (prov or "proveedor") if ch.isalnum() or ch in " _-")[:30].strip() or "proveedor"
    return send_file(out, as_attachment=True, download_name=f"HOJA_{safe}_{hoy()}.pdf")


# ----------------------------------------------------------------------------
# Exportar a Excel
# ----------------------------------------------------------------------------
@app.route("/api/exportar")
def api_exportar():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    db = get_db()
    rows = db.execute("SELECT * FROM actividades ORDER BY id").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AVANCE"
    cols = ["ID", "Código", "Bloque", "Área", "Giro", "Proveedor", "Partida",
            "Tipo", "¿Aplica?", "% Avance", "F. Inicio", "F. Fin",
            "Duración (días)", "Estatus", "Notas"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for i, r in enumerate(rows, 2):
        vals = [r["id"], r["codigo"], r["bloque"], r["area"], r["giro"],
                r["proveedor"], r["partida"], r["tipo"], r["aplica"],
                r["avance"], r["f_inicio"], r["f_fin"], r["duracion_dias"],
                r["estatus"], r["notas"]]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
    widths = [6, 10, 20, 24, 14, 22, 38, 12, 9, 9, 12, 12, 14, 14, 30]
    from openpyxl.utils import get_column_letter
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    out = os.path.join(BASE_DIR, "data", "export_avance.xlsx")
    wb.save(out)
    return send_file(out, as_attachment=True,
                     download_name=f"AVANCE_HAP_{hoy()}.xlsx")


# ----------------------------------------------------------------------------
# Plantilla de captura para llenar rápido (con desplegables) y reimportar
# ----------------------------------------------------------------------------
@app.route("/api/plantilla_captura.xlsx")
def api_plantilla_captura():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    rows = _filtrar_actividades(request.args)
    hoy_d = datetime.date.today()
    # causas para el desplegable
    dbc = get_db()
    causas = [r[0] for r in dbc.execute("SELECT nombre FROM causas ORDER BY base DESC,nombre").fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Captura"

    ws["A1"] = "PLANTILLA DE CAPTURA DE AVANCE — HAP"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:H1")
    ws["A2"] = ("Llena las columnas AMARILLAS. NO cambies la columna Código. "
                "Al terminar, súbela en la plataforma con 'Importar avances'.")
    ws["A2"].font = Font(size=9, italic=True, color="B00020")
    ws.merge_cells("A2:H2")

    # Encabezados: las primeras 4 son de solo lectura (referencia); las 3 amarillas se llenan
    cols = ["Código", "Área", "Partida", "Avance actual",
            "NUEVO avance %", "Fecha compromiso (AAAA-MM-DD)", "Causa de retraso"]
    hdr = 4
    azul = PatternFill("solid", fgColor="1F4E78")
    amar = PatternFill("solid", fgColor="FFF2CC")
    gris = PatternFill("solid", fgColor="EDEDED")
    thin = Side(style="thin", color="CCCCCC")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=hdr, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = azul
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = bd
    ws.row_dimensions[hdr].height = 30

    i = hdr + 1
    for r in rows:
        ws.cell(row=i, column=1, value=r["codigo"]).border = bd
        ws.cell(row=i, column=2, value=r["area"]).border = bd
        ws.cell(row=i, column=3, value=r["partida"]).border = bd
        c4 = ws.cell(row=i, column=4, value=f"{r['avance'] or 0}%"); c4.border = bd; c4.fill = gris
        # columnas amarillas para llenar
        for col in (5, 6, 7):
            cc = ws.cell(row=i, column=col); cc.border = bd; cc.fill = amar
        i += 1
    ultima = i - 1

    # Validaciones: avance por lista 0/25/50/75/100, causa por lista de causas
    if ultima >= hdr + 1:
        dv_av = DataValidation(type="list", formula1='"0,25,50,75,100"', allow_blank=True)
        ws.add_data_validation(dv_av)
        dv_av.add(f"E{hdr+1}:E{ultima}")
        # causas: si son pocas caben en formula directa; si no, usar hoja aparte
        causas_txt = ",".join(c.replace(",", " ") for c in causas)
        if len(causas_txt) < 240:
            dv_ca = DataValidation(type="list", formula1=f'"{causas_txt}"', allow_blank=True)
            ws.add_data_validation(dv_ca)
            dv_ca.add(f"G{hdr+1}:G{ultima}")
        else:
            ws2 = wb.create_sheet("causas")
            for k, c in enumerate(causas, 1):
                ws2.cell(row=k, column=1, value=c)
            dv_ca = DataValidation(type="list",
                                   formula1=f"causas!$A$1:$A${len(causas)}", allow_blank=True)
            ws.add_data_validation(dv_ca)
            dv_ca.add(f"G{hdr+1}:G{ultima}")

    anchos = [11, 22, 42, 12, 14, 20, 26]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"
    out = os.path.join(BASE_DIR, "data", "plantilla_captura.xlsx")
    wb.save(out)
    prov = request.args.get("proveedor", "")
    safe = "".join(ch for ch in (prov or "todos") if ch.isalnum() or ch in " _-")[:30].strip() or "todos"
    return send_file(out, as_attachment=True, download_name=f"CAPTURA_{safe}_{hoy()}.xlsx")


@app.route("/api/importar_avances", methods=["POST"])
def api_importar_avances():
    """Lee un Excel con Código + NUEVO avance / fecha / causa y actualiza."""
    import openpyxl
    if "archivo" not in request.files:
        return jsonify({"error": "No llegó ningún archivo"}), 400
    f = request.files["archivo"]
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el Excel: {e}"}), 400
    ws = wb["Captura"] if "Captura" in wb.sheetnames else wb.active

    # localizar la fila de encabezados y las columnas por su nombre
    encabezados = {}
    fila_hdr = None
    for ri in range(1, min(12, ws.max_row + 1)):
        vals = {(ws.cell(row=ri, column=ci).value or "").__str__().strip().lower(): ci
                for ci in range(1, ws.max_column + 1)}
        # la fila de encabezados tiene una celda que es EXACTAMENTE "código"/"codigo"
        if "código" in vals or "codigo" in vals:
            encabezados = vals; fila_hdr = ri; break
    if not fila_hdr:
        return jsonify({"error": "No encontré la fila de encabezados con la columna 'Código'"}), 400

    def col_de(*claves):
        for k, ci in encabezados.items():
            for clave in claves:
                if clave in k:
                    return ci
        return None
    c_cod = encabezados.get("código") or encabezados.get("codigo")
    c_av = col_de("nuevo avance", "nuevo")
    c_fe = col_de("fecha compromiso", "compromiso")
    c_ca = col_de("causa")

    db = get_db()
    actualizadas, sin_match, errores = 0, [], 0
    for ri in range(fila_hdr + 1, ws.max_row + 1):
        codigo = ws.cell(row=ri, column=c_cod).value
        if not codigo:
            continue
        codigo = str(codigo).strip()
        act = db.execute("SELECT * FROM actividades WHERE codigo=?", (codigo,)).fetchone()
        if not act:
            sin_match.append(codigo); continue
        cambios, args = [], []
        # avance
        if c_av:
            v = ws.cell(row=ri, column=c_av).value
            if v is not None and str(v).strip() != "":
                try:
                    av = int(float(str(v).replace("%", "").strip()))
                    av = max(0, min(100, av))
                    if av != (act["avance"] or 0):
                        cambios.append("avance=?"); args.append(av)
                        cambios.append("estatus=?"); args.append(estatus_por_avance(av))
                except ValueError:
                    errores += 1
        # fecha compromiso
        if c_fe:
            v = ws.cell(row=ri, column=c_fe).value
            if v is not None and str(v).strip() != "":
                d = parse_date(str(v)[:10]) if not isinstance(v, datetime.datetime) else v.date()
                if d:
                    iso = d.isoformat()
                    if iso != (act["f_fin"] or ""):
                        cambios.append("f_fin=?"); args.append(iso)
        # causa
        if c_ca:
            v = ws.cell(row=ri, column=c_ca).value
            if v is not None and str(v).strip() != "":
                causa = str(v).strip()
                if causa != (act["causa_retraso"] or ""):
                    cambios.append("causa_retraso=?"); args.append(causa)
                    cambios.append("causa_por=?"); args.append("Importado de Excel")
                    cambios.append("causa_fecha=?"); args.append(datetime.date.today().isoformat())
        if cambios:
            cambios.append("actualizado=?")
            args.append(datetime.datetime.now().isoformat(timespec="seconds"))
            args.append(act["id"])
            db.execute(f"UPDATE actividades SET {','.join(cambios)} WHERE id=?", args)
            actualizadas += 1
    db.commit()
    return jsonify({
        "ok": True, "actualizadas": actualizadas,
        "sin_coincidencia": sin_match, "errores": errores,
    })


# ----------------------------------------------------------------------------
# Autenticación y portal de proveedores
# ----------------------------------------------------------------------------
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json()
    usuario = (data.get("usuario") or "").strip()
    clave = data.get("clave") or ""
    db = get_db()
    u = db.execute("SELECT * FROM usuarios WHERE usuario=?", (usuario,)).fetchone()
    if not u or u["clave"] != hash_clave(clave):
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401
    # candado: usuario desactivado no puede entrar
    if "activo" in u.keys() and u["activo"] == 0:
        return jsonify({"error": "Tu acceso está desactivado. Contacta al administrador."}), 403
    # contar el acceso
    try:
        db.execute("UPDATE usuarios SET num_logins=COALESCE(num_logins,0)+1, ultimo_login=? WHERE id=?",
                   (datetime.datetime.now().isoformat(timespec="seconds"), u["id"]))
        db.commit()
    except Exception:
        pass
    session["usuario"] = u["usuario"]
    session["rol"] = u["rol"]
    session["proveedor"] = u["proveedor"]
    session["mundo"] = u["mundo"] if "mundo" in u.keys() else "obra"
    return jsonify({"ok": True, "usuario": u["usuario"], "rol": u["rol"],
                    "proveedor": u["proveedor"], "mundo": session["mundo"]})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/quien_soy")
def api_quien_soy():
    if "usuario" not in session:
        return jsonify({"login": False})
    return jsonify({"login": True, "usuario": session["usuario"],
                    "rol": session["rol"], "proveedor": session.get("proveedor"),
                    "mundo": session.get("mundo", "obra")})


# --- Gestión de usuarios (solo admin) ---
@app.route("/api/usuarios", methods=["GET"])
@requiere_admin
def api_usuarios():
    db = get_db()
    rows = db.execute("SELECT id,usuario,proveedor,rol,mundo,clave_cambiada,creado,activo,num_logins,ultimo_login,telefono FROM usuarios ORDER BY rol,mundo,usuario").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/usuarios", methods=["POST"])
@requiere_admin
def api_crear_usuario():
    db = get_db()
    data = request.get_json()
    usuario = (data.get("usuario") or "").strip()
    clave = data.get("clave") or ""
    proveedor = data.get("proveedor") or None
    mundo = data.get("mundo", "obra")
    telefono = (data.get("telefono") or "").strip() or None
    # rol: por defecto proveedor. El admin puede crear supervisores.
    rol = data.get("rol", "proveedor")
    if rol not in ("proveedor", "supervisor_obra", "supervisor_depto"):
        rol = "proveedor"
    # el supervisor de obra vive en mundo 'obra'; el de depto en 'interno'
    if rol == "supervisor_obra":
        mundo = "obra"
    elif rol == "supervisor_depto":
        mundo = "interno"
    if not usuario or not clave:
        return jsonify({"error": "Usuario y contraseña son obligatorios"}), 400
    try:
        db.execute(
            "INSERT INTO usuarios (usuario,clave,proveedor,rol,mundo,telefono,activo,num_logins,creado) VALUES (?,?,?,?,?,?,1,0,?)",
            (usuario, hash_clave(clave), proveedor, rol, mundo, telefono,
             datetime.datetime.now().isoformat(timespec="seconds")))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({"error": "Ese usuario ya existe"}), 400
    return jsonify({"ok": True})


@app.route("/api/usuario/<int:uid>", methods=["DELETE"])
@requiere_admin
def api_borrar_usuario(uid):
    db = get_db()
    u = db.execute("SELECT rol FROM usuarios WHERE id=?", (uid,)).fetchone()
    if u and u["rol"] == "admin":
        return jsonify({"error": "No se puede borrar al administrador"}), 400
    db.execute("DELETE FROM usuarios WHERE id=?", (uid,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/usuario/<int:uid>/clave", methods=["POST"])
@requiere_admin
def api_cambiar_clave(uid):
    db = get_db()
    data = request.get_json()
    nueva = data.get("clave") or ""
    if not nueva:
        return jsonify({"error": "Escribe la nueva contraseña"}), 400
    db.execute("UPDATE usuarios SET clave=? WHERE id=?", (hash_clave(nueva), uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/usuario/<int:uid>/activo", methods=["POST"])
@requiere_admin
def api_usuario_activo(uid):
    """Activa o desactiva el acceso de un usuario (sin borrarlo)."""
    db = get_db()
    data = request.get_json() or {}
    activo = 1 if data.get("activo") else 0
    u = db.execute("SELECT rol FROM usuarios WHERE id=?", (uid,)).fetchone()
    if u and u["rol"] == "admin":
        return jsonify({"error": "No se puede desactivar al administrador"}), 400
    db.execute("UPDATE usuarios SET activo=? WHERE id=?", (activo, uid))
    db.commit()
    return jsonify({"ok": True, "activo": activo})


@app.route("/api/usuario/<int:uid>/telefono", methods=["POST"])
@requiere_admin
def api_usuario_telefono(uid):
    """Guarda o actualiza el telefono (para el aviso por WhatsApp)."""
    db = get_db()
    data = request.get_json() or {}
    tel = (data.get("telefono") or "").strip() or None
    db.execute("UPDATE usuarios SET telefono=? WHERE id=?", (tel, uid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/usuario/<int:uid>/whatsapp", methods=["GET"])
@requiere_admin
def api_usuario_whatsapp(uid):
    """Arma el link de WhatsApp con un mensaje listo avisando de nuevas asignaciones.
    No envia nada: devuelve el link para que el admin le de click."""
    db = get_db()
    u = db.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not u:
        return jsonify({"error": "Usuario no existe"}), 404
    tel = (u["telefono"] or "").strip() if "telefono" in u.keys() else ""
    if not tel:
        return jsonify({"error": "Este usuario no tiene telefono registrado"}), 400
    # contar nuevas por reconocer de ese proveedor
    proveedor = u["proveedor"]
    mundo = u["mundo"] if "mundo" in u.keys() else "obra"
    col = "departamento" if mundo == "interno" else "proveedor"
    n = db.execute(
        f"SELECT COUNT(*) FROM actividades WHERE {col}=? AND reconocida='NO' "
        "AND (estado_val IS NULL OR estado_val='validado') "
        "AND (mundo=? OR (mundo IS NULL AND ?='obra'))",
        (proveedor, mundo, mundo)).fetchone()[0]
    # normalizar telefono: solo digitos; si no trae lada pais, anteponer 52 (Mexico)
    solo_num = "".join(ch for ch in tel if ch.isdigit())
    if len(solo_num) == 10:
        solo_num = "52" + solo_num
    liga = request.host_url.rstrip("/") + "/portal"
    if n > 0:
        texto = (f"Hola {proveedor or u['usuario']}, tienes {n} "
                 f"{'nueva actividad asignada' if n == 1 else 'nuevas actividades asignadas'} "
                 f"en el Control de Obra HAP. Entra a reconocerlas y reportar tu avance: {liga}")
    else:
        texto = (f"Hola {proveedor or u['usuario']}, te recuerdo revisar tus actividades "
                 f"en el Control de Obra HAP: {liga}")
    import urllib.parse
    url = "https://wa.me/" + solo_num + "?text=" + urllib.parse.quote(texto)
    return jsonify({"ok": True, "url": url, "mensaje": texto, "nuevas": n, "telefono": tel})


# --- Portal del proveedor: ve y reporta SOLO lo suyo ---
@app.route("/portal")
def portal_page():
    return render_template("portal.html")


@app.route("/api/portal/mapa")
@requiere_login
def api_portal_mapa():
    """Mapa bloque->áreas para el formulario de nueva actividad.
    Externo: solo sus bloques. Interno: todo el mapa de la obra (su labor cruza el hospital)."""
    _, rol, proveedor = usuario_actual()
    db = get_db()
    col, mundo = col_duenio()
    mapa = {}
    if mundo == "interno":
        # todo el mapa de la obra
        filas = db.execute(
            "SELECT DISTINCT bloque, area FROM actividades "
            "WHERE bloque IS NOT NULL AND bloque<>'' AND area IS NOT NULL AND area<>'' "
            "ORDER BY bloque, area").fetchall()
        # bloques donde el depto YA tiene trabajo (para marcar lo que es 'fuera de zona')
        suyos = [r[0] for r in db.execute(
            f"SELECT DISTINCT bloque FROM actividades WHERE {col}=? AND bloque IS NOT NULL",
            (proveedor,)).fetchall()]
    else:
        filas = db.execute(
            f"SELECT DISTINCT bloque, area FROM actividades WHERE {col}=? "
            "AND bloque IS NOT NULL AND bloque<>'' AND area IS NOT NULL AND area<>'' "
            "ORDER BY bloque, area", (proveedor,)).fetchall()
        suyos = None  # externo: todos los que ve ya son suyos
    for r in filas:
        mapa.setdefault(r["bloque"], []).append(r["area"])
    return jsonify({"mapa": mapa, "mundo": mundo, "bloques_propios": suyos})


@app.route("/api/cambiar_mi_clave", methods=["POST"])
@requiere_login
def api_cambiar_mi_clave():
    usuario, _, _ = usuario_actual()
    db = get_db()
    data = request.get_json() or {}
    clave_actual = data.get("clave_actual") or ""
    clave_nueva = data.get("clave_nueva") or ""
    if not clave_actual or not clave_nueva:
        return jsonify({"error": "Ingresa tu contraseña actual y la nueva contraseña"}), 400
    if len(clave_nueva) < 4:
        return jsonify({"error": "La nueva contraseña debe tener al menos 4 caracteres"}), 400
    
    u = db.execute("SELECT * FROM usuarios WHERE usuario=?", (usuario,)).fetchone()
    if not u or u["clave"] != hash_clave(clave_actual):
        return jsonify({"error": "La contraseña actual es incorrecta"}), 400
        
    db.execute("UPDATE usuarios SET clave=?, clave_cambiada=1 WHERE usuario=?",
               (hash_clave(clave_nueva), usuario))
    db.commit()
    return jsonify({"ok": True, "mensaje": "Contraseña actualizada exitosamente"})


@app.route("/api/portal/avance_zona", methods=["GET"])
@requiere_login
def api_portal_avance_zona():
    """Vista de solo lectura del avance de todos los gremios en la zona del proveedor."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    col, mundo = col_duenio()
    
    bloque = request.args.get("bloque")
    area = request.args.get("area")
    
    # Determinar qué bloques/áreas puede consultar
    if mundo == "interno":
        # Interno: ve todo el hospital
        cond = ["(aplica IS NULL OR aplica<>'NO')", "(estado_val IS NULL OR estado_val='validado')"]
        args = []
        bloques_disp = [r[0] for r in db.execute("SELECT DISTINCT bloque FROM actividades WHERE bloque IS NOT NULL AND bloque<>'' ORDER BY bloque").fetchall()]
    else:
        # Externo: ve los bloques donde tiene partidas asignadas
        bloques_suyos = [r[0] for r in db.execute(
            f"SELECT DISTINCT bloque FROM actividades WHERE {col}=? AND bloque IS NOT NULL AND bloque<>''",
            (proveedor,)).fetchall()]
        if not bloques_suyos:
            return jsonify({"bloques_disponibles": [], "actividades": []})
        marcas = ",".join("?" * len(bloques_suyos))
        cond = [f"bloque IN ({marcas})", "(aplica IS NULL OR aplica<>'NO')", "(estado_val IS NULL OR estado_val='validado')"]
        args = list(bloques_suyos)
        bloques_disp = sorted(bloques_suyos)
        
    if bloque:
        cond.append("bloque = ?"); args.append(bloque)
    if area:
        cond.append("area = ?"); args.append(area)
        
    q = f"SELECT id, codigo, bloque, area, giro, proveedor, departamento, mundo, partida, tipo, tipo_partida, avance, estatus, f_inicio, f_fin FROM actividades WHERE {' AND '.join(cond)} ORDER BY bloque, area, id"
    filas = db.execute(q, args).fetchall()
    
    return jsonify({
        "bloques_disponibles": bloques_disp,
        "actividades": [dict(r) for r in filas]
    })


@app.route("/api/portal/mis_actividades")
@requiere_login
def api_portal_mis():
    _, rol, proveedor = usuario_actual()
    if es_gestor(rol):
        return jsonify({"error": "Los gestores usan el panel principal"}), 400
    db = get_db()
    col, mundo = col_duenio()
    rows = db.execute(
        f"SELECT * FROM actividades WHERE {col}=? AND (mundo=? OR (mundo IS NULL AND ?='obra')) "
        "AND (reconocida IS NULL OR reconocida<>'RECHAZADA') "
        "ORDER BY estado_val DESC, area, id",
        (proveedor, mundo, mundo)).fetchall()
    
    salida = []
    for r in rows:
        d = dict(r)
        dep_st = evaluar_dependencias_area(db, r["area"], r["tipo_partida"] or r["giro"])
        d["dep_bloqueada"] = dep_st["bloqueada"]
        d["dep_estado"] = dep_st["estado"]
        d["dep_detalle"] = dep_st["detalle"]
        salida.append(d)
    return jsonify(salida)


@app.route("/api/portal/reportar_avance/<int:aid>", methods=["POST"])
@requiere_login
def api_portal_reportar(aid):
    """El proveedor declara avance en una partida oficial suya. Entra directo como 'declarado'."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    col, _ = col_duenio()
    if not a or a[col] != proveedor:
        return jsonify({"error": "Esa partida no es tuya"}), 403
    # candado: no puede reportar avance si no ha reconocido la actividad
    if a["reconocida"] != "SÍ":
        return jsonify({"error": "Primero tienes que reconocer esta actividad"}), 400
    data = request.get_json() or {}
    campos, args = [], []
    if "avance_decl" in data:
        av = max(0, min(100, int(data.get("avance_decl") or 0)))
        registrar_historial(db, aid, "avance reportado", a["avance_decl"], av, usuario)
        campos += ["avance_decl=?", "avance_decl_por=?", "avance_decl_fecha=?", "avance_decl_rechazado=0", "rechazo_motivo=NULL"]
        args += [av, usuario, datetime.date.today().isoformat()]
    for c in ("f_inicio", "f_fin", "definido", "definido_por", "nota_proveedor"):
        if c in data:
            registrar_historial(db, aid, c, a[c], data[c], usuario)
            campos.append(f"{c}=?"); args.append(data[c])
    if campos:
        campos.append("actualizado=?"); args.append(datetime.datetime.now().isoformat(timespec="seconds"))
        args.append(aid)
        db.execute(f"UPDATE actividades SET {','.join(campos)} WHERE id=?", args)
        # aviso para el admin
        db.execute("INSERT INTO avisos (fecha,proveedor,usuario,tipo,detalle,visto) VALUES (?,?,?,?,?,0)",
                   (datetime.datetime.now().isoformat(timespec="seconds"), proveedor, usuario,
                    "avance", f"Reportó avance en {a['codigo']} · {(a['partida'] or '')[:50]}"))
        db.commit()
    return jsonify({"ok": True})


@app.route("/api/portal/reconocer/<int:aid>", methods=["POST"])
@requiere_login
def api_portal_reconocer(aid):
    """El proveedor acepta que una partida es su trabajo. Sin esto no puede reportar avance."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    col, _ = col_duenio()
    if not a or a[col] != proveedor:
        return jsonify({"error": "Esa partida no es tuya"}), 403
    registrar_historial(db, aid, "reconocimiento", "sin reconocer", "reconocida", usuario)
    db.execute("UPDATE actividades SET reconocida='SÍ', reconocida_por=?, reconocida_fecha=?, "
               "no_reconocida_nota=NULL WHERE id=?",
               (usuario, datetime.date.today().isoformat(), aid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/portal/reconocer_bloque", methods=["POST"])
@requiere_login
def api_portal_reconocer_bloque():
    """Reconoce en lote todas las pendientes de un bloque (o todas si no se manda bloque)."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    data = request.get_json() or {}
    bloque = data.get("bloque")
    col, mundo = col_duenio()
    cond = f"{col}=? AND reconocida='NO' AND (estado_val IS NULL OR estado_val='validado') AND (mundo=? OR (mundo IS NULL AND ?='obra'))"
    args = [proveedor, mundo, mundo]
    if bloque:
        cond += " AND bloque=?"; args.append(bloque)
    filas = db.execute(f"SELECT id FROM actividades WHERE {cond}", args).fetchall()
    for f in filas:
        registrar_historial(db, f["id"], "reconocimiento", "sin reconocer", "reconocida (en lote)", usuario)
    db.execute(f"UPDATE actividades SET reconocida='SÍ', reconocida_por=?, reconocida_fecha=? WHERE {cond}",
               [usuario, datetime.date.today().isoformat()] + args)
    db.commit()
    return jsonify({"ok": True, "reconocidas": len(filas)})


@app.route("/api/portal/no_reconozco/<int:aid>", methods=["POST"])
@requiere_login
def api_portal_no_reconozco(aid):
    """El proveedor dice que esa partida NO es suya. Te avisa a ti para corregir."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    col, _ = col_duenio()
    if not a or a[col] != proveedor:
        return jsonify({"error": "Esa partida no es tuya"}), 403
    data = request.get_json() or {}
    nota = (data.get("nota") or "").strip()
    registrar_historial(db, aid, "no reconocida", "", nota or "el proveedor no reconoce esta actividad", usuario)
    db.execute("UPDATE actividades SET no_reconocida_nota=?, reconocida='RECHAZADA' WHERE id=?", (nota or "No reconocida", aid))
    db.execute("INSERT INTO avisos (fecha,proveedor,usuario,tipo,detalle,visto) VALUES (?,?,?,?,?,0)",
               (datetime.datetime.now().isoformat(timespec="seconds"), proveedor, usuario,
                "no_reconocida", f"No reconoce {a['codigo']}: {(a['partida'] or '')[:50]}"))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/portal/pendientes_reconocer")
@requiere_login
def api_portal_pend_reconocer():
    """Conteo de lo que el proveedor tiene sin reconocer (para el botón de 'nuevas asignadas')."""
    _, rol, proveedor = usuario_actual()
    db = get_db()
    col, mundo = col_duenio()
    n = db.execute(f"SELECT COUNT(*) FROM actividades WHERE {col}=? AND reconocida='NO' "
                   "AND (estado_val IS NULL OR estado_val='validado') "
                   "AND (mundo=? OR (mundo IS NULL AND ?='obra'))",
                   (proveedor, mundo, mundo)).fetchone()[0]
    return jsonify({"pendientes": n})


@app.route("/api/portal/nueva", methods=["POST"])
@requiere_login
def api_portal_nueva():
    """El proveedor propone una tarea NUEVA. Queda en estado 'propuesta' hasta que el admin valide."""
    usuario, rol, proveedor = usuario_actual()
    db = get_db()
    data = request.get_json()
    partida = (data.get("partida") or "").strip()
    if not partida:
        return jsonify({"error": "Escribe qué actividad es"}), 400
    col, mundo = col_duenio()
    # detectar ANTES de insertar si el interno propone en un bloque donde NO tenía trabajo
    fuera_zona = 0
    if mundo == "interno" and data.get("bloque"):
        tiene = db.execute(
            f"SELECT COUNT(*) FROM actividades WHERE {col}=? AND bloque=?",
            (proveedor, data.get("bloque"))).fetchone()[0]
        fuera_zona = 1 if tiene == 0 else 0
    # el giro y tipo se ponen solos según el proveedor/departamento (los más comunes de sus partidas)
    ref = db.execute(
        f"SELECT giro, tipo_partida, tipo_interno FROM actividades WHERE {col}=? "
        "GROUP BY giro ORDER BY COUNT(*) DESC LIMIT 1", (proveedor,)).fetchone()
    giro = (ref["giro"] if ref else "") or data.get("giro", "")
    tipo_partida = (ref["tipo_partida"] if ref else "Construcción")
    tipo_interno = data.get("tipo_interno") or (ref["tipo_interno"] if ref else None)
    # columnas dueñas: en interno se llena departamento; en obra, proveedor
    prov_val = proveedor if mundo == "obra" else None
    depto_val = proveedor if mundo == "interno" else None
    # código provisional
    codigo = "PROP-" + datetime.datetime.now().strftime("%m%d%H%M%S")
    cur = db.execute(
        """INSERT INTO actividades
        (codigo,bloque,area,giro,proveedor,departamento,mundo,tipo_interno,partida,tipo_partida,aplica,avance,
         avance_decl,f_inicio,f_fin,estatus,definido,definido_por,nota_proveedor,
         origen,estado_val,creado_por,reconocida,reconocida_por,reconocida_fecha,fuera_zona,actualizado)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (codigo, data.get("bloque"), data.get("area"), giro, prov_val, depto_val, mundo, tipo_interno, partida,
         tipo_partida, "SÍ", 0, int(data.get("avance_decl") or 0),
         data.get("f_inicio"), data.get("f_fin"), "Pendiente",
         data.get("definido", "NO"), data.get("definido_por"), data.get("nota_proveedor"),
         "propuesta", "propuesta", usuario,
         "SÍ", usuario, datetime.date.today().isoformat(), fuera_zona,
         datetime.datetime.now().isoformat(timespec="seconds")))
    detalle = f"Propuso nueva actividad: {partida[:60]}"
    if fuera_zona:
        detalle = f"⚠️ FUERA DE SU ZONA ({data.get('bloque')}): {partida[:50]}"
    db.execute("INSERT INTO avisos (fecha,proveedor,usuario,tipo,detalle,visto) VALUES (?,?,?,?,?,0)",
               (datetime.datetime.now().isoformat(timespec="seconds"), proveedor, usuario,
                "propuesta", detalle))
    db.commit()
    return jsonify({"ok": True, "codigo": codigo, "fuera_zona": bool(fuera_zona)})


@app.route("/api/validacion/atencion")
@requiere_gestor
def api_val_atencion():
    """Actividades que necesitan tu atención: rechazadas y las que un proveedor no reconoce."""
    db = get_db()
    rechazadas = db.execute(
        "SELECT * FROM actividades WHERE estado_val='rechazada' OR avance_decl_rechazado=1 ORDER BY proveedor, actualizado DESC").fetchall()
    no_recon = db.execute(
        "SELECT * FROM actividades WHERE no_reconocida_nota IS NOT NULL AND no_reconocida_nota<>'' "
        "AND reconocida<>'SÍ' ORDER BY proveedor, actualizado DESC").fetchall()
    return jsonify({
        "rechazadas": [dict(r) for r in rechazadas],
        "no_reconocidas": [dict(r) for r in no_recon],
    })


@app.route("/api/validacion/reactivar/<int:aid>", methods=["POST"])
@requiere_gestor
def api_val_reactivar(aid):
    """Reactiva una actividad rechazada: vuelve a propuesta para revisarla de nuevo."""
    db = get_db()
    quien_admin = session.get("usuario", "admin")
    registrar_historial(db, aid, "actividad propuesta", "rechazada", "reactivada", quien_admin)
    db.execute("UPDATE actividades SET estado_val='propuesta', avance_decl_rechazado=0, actualizado=? WHERE id=?",
               (datetime.datetime.now().isoformat(timespec="seconds"), aid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/validacion/limpiar_no_reconocida/<int:aid>", methods=["POST"])
@requiere_gestor
def api_val_limpiar_noreco(aid):
    """Después de corregir una actividad que el proveedor no reconocía, limpias la marca."""
    db = get_db()
    db.execute("UPDATE actividades SET no_reconocida_nota=NULL, reconocida='NO', reconocida_por=NULL, reconocida_fecha=NULL WHERE id=?", (aid,))
    db.commit()
    return jsonify({"ok": True})


# --- Validación (admin) ---
@app.route("/api/validacion/pendientes")
@requiere_gestor
def api_val_pendientes():
    db = get_db()
    props = db.execute(
        "SELECT * FROM actividades WHERE estado_val='propuesta' ORDER BY proveedor, actualizado DESC").fetchall()
    # partidas oficiales con avance declarado distinto al validado
    decl = db.execute(
        "SELECT * FROM actividades WHERE origen='oficial' AND avance_decl IS NOT NULL "
        "AND avance_decl <> avance ORDER BY proveedor, actualizado DESC").fetchall()
    return jsonify({
        "propuestas": [dict(r) for r in props],
        "avances": [dict(r) for r in decl],
    })


@app.route("/api/validacion/aviso_conteo")
@requiere_gestor
def api_val_conteo():
    db = get_db()
    p = db.execute("SELECT COUNT(*) FROM actividades WHERE estado_val='propuesta'").fetchone()[0]
    a = db.execute("SELECT COUNT(*) FROM actividades WHERE origen='oficial' "
                   "AND avance_decl IS NOT NULL AND avance_decl <> avance").fetchone()[0]
    return jsonify({"propuestas": p, "avances": a, "total": p + a})


@app.route("/api/validacion/aprobar/<int:aid>", methods=["POST"])
@requiere_gestor
def api_val_aprobar(aid):
    db = get_db()
    a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify({"error": "no existe"}), 404
    data = request.get_json() or {}
    quien_admin = session.get("usuario", "admin")
    if a["estado_val"] == "propuesta":
        # aprobar la tarea nueva: pasa a oficial y validada
        av = a["avance_decl"] if a["avance_decl"] is not None else 0
        registrar_historial(db, aid, "actividad propuesta", "propuesta", "aprobada y validada", quien_admin)
        db.execute("UPDATE actividades SET origen='oficial', estado_val='validado', "
                   "avance=?, estatus=?, avance_decl_rechazado=0, rechazo_motivo=NULL, actualizado=? WHERE id=?",
                   (av, estatus_por_avance(av),
                    datetime.datetime.now().isoformat(timespec="seconds"), aid))
    else:
        # firmar el avance declarado: el avance oficial toma el valor declarado
        av = a["avance_decl"] if a["avance_decl"] is not None else a["avance"]
        registrar_historial(db, aid, "avance validado", a["avance"], av, quien_admin)
        db.execute("UPDATE actividades SET avance=?, estatus=?, avance_decl_rechazado=0, rechazo_motivo=NULL, actualizado=? WHERE id=?",
                   (av, estatus_por_avance(av),
                    datetime.datetime.now().isoformat(timespec="seconds"), aid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/validacion/rechazar/<int:aid>", methods=["POST"])
@requiere_gestor
def api_val_rechazar(aid):
    db = get_db()
    a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify({"error": "no existe"}), 404
    data = request.get_json() or {}
    motivo = (data.get("motivo") or "Rechazado por administración").strip()
    quien_admin = session.get("usuario", "admin")
    fecha_hoy = datetime.date.today().isoformat()
    
    if a["estado_val"] == "propuesta":
        # rechazar tarea nueva: se marca rechazada con motivo
        registrar_historial(db, aid, "actividad propuesta", "propuesta", f"rechazada: {motivo}", quien_admin)
        db.execute("UPDATE actividades SET estado_val='rechazada', rechazo_motivo=?, rechazado_por=?, "
                   "rechazado_fecha=?, actualizado=? WHERE id=?",
                   (motivo, quien_admin, fecha_hoy, datetime.datetime.now().isoformat(timespec="seconds"), aid))
    else:
        # rechazar avance declarado: se guarda el rechazo con su motivo para que el proveedor se entere en su portal
        registrar_historial(db, aid, "avance reportado", a["avance_decl"], f"rechazado por admin ({motivo})", quien_admin)
        db.execute("UPDATE actividades SET avance_decl_rechazado=1, rechazo_motivo=?, rechazado_por=?, "
                   "rechazado_fecha=?, avance_decl=NULL, avance_decl_por=NULL, avance_decl_fecha=NULL, "
                   "actualizado=? WHERE id=?",
                   (motivo, quien_admin, fecha_hoy, datetime.datetime.now().isoformat(timespec="seconds"), aid))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/validacion/aprobar_todas", methods=["POST"])
@requiere_gestor
def api_val_aprobar_todas():
    db = get_db()
    data = request.get_json() or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "sin ids"}), 400
    n = 0
    quien_admin = session.get("usuario", "admin")
    for aid in ids:
        a = db.execute("SELECT * FROM actividades WHERE id=?", (aid,)).fetchone()
        if not a:
            continue
        av = a["avance_decl"] if a["avance_decl"] is not None else a["avance"]
        if a["estado_val"] == "propuesta":
            registrar_historial(db, aid, "actividad propuesta", "propuesta", "aprobada y validada", quien_admin)
            db.execute("UPDATE actividades SET origen='oficial', estado_val='validado', "
                       "avance=?, estatus=?, avance_decl_rechazado=0, rechazo_motivo=NULL, actualizado=? WHERE id=?",
                       (av, estatus_por_avance(av),
                        datetime.datetime.now().isoformat(timespec="seconds"), aid))
        else:
            registrar_historial(db, aid, "avance validado", a["avance"], av, quien_admin)
            db.execute("UPDATE actividades SET avance=?, estatus=?, avance_decl_rechazado=0, rechazo_motivo=NULL, actualizado=? WHERE id=?",
                       (av, estatus_por_avance(av),
                        datetime.datetime.now().isoformat(timespec="seconds"), aid))
        n += 1
    db.commit()
    return jsonify({"ok": True, "aprobadas": n})


# ----------------------------------------------------------------------------
# API — Dependencias (Modelo por área)
# ----------------------------------------------------------------------------
@app.route("/api/dependencias", methods=["GET"])
@requiere_login
def api_dependencias_list():
    db = get_db()
    area = request.args.get("area")
    q = "SELECT * FROM dependencias"
    args = []
    if area:
        q += " WHERE area=?"; args.append(area)
    q += " ORDER BY area, id"
    rows = db.execute(q, args).fetchall()
    
    salida = []
    for r in rows:
        d = dict(r)
        st = evaluar_dependencias_area(db, r["area"], r["tipo_sucesor"])
        d["evaluacion"] = st
        salida.append(d)
    return jsonify(salida)


@app.route("/api/dependencias", methods=["POST"])
@requiere_gestor
def api_dependencias_crear():
    db = get_db()
    data = request.get_json() or {}
    area = (data.get("area") or "").strip()
    sucesor = (data.get("tipo_sucesor") or "").strip()
    predecesores = (data.get("tipos_predecesores") or "").strip()
    umbral = max(1, min(100, int(data.get("umbral") or 100)))
    paralelo = 1 if data.get("permite_paralelo") else 0
    
    if not area or not sucesor or not predecesores:
        return jsonify({"error": "Área, tipo sucesor y tipos predecesores son obligatorios"}), 400
        
    db.execute(
        "INSERT INTO dependencias (area, tipo_sucesor, tipos_predecesores, umbral, permite_paralelo, creado) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (area, sucesor, predecesores, umbral, paralelo, datetime.datetime.now().isoformat(timespec="seconds"))
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/dependencias/<int:did>", methods=["DELETE"])
@requiere_gestor
def api_dependencias_borrar(did):
    db = get_db()
    db.execute("DELETE FROM dependencias WHERE id=?", (did,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/dependencias/<int:did>/forzar_liberacion", methods=["POST"])
@requiere_gestor
def api_dependencias_forzar(did):
    db = get_db()
    data = request.get_json() or {}
    nota = (data.get("nota") or "").strip()
    if not nota:
        return jsonify({"error": "Se requiere una nota justificando la liberación anticipada"}), 400
        
    admin_user = session.get("usuario", "admin")
    fecha_hoy = datetime.date.today().isoformat()
    
    dep = db.execute("SELECT * FROM dependencias WHERE id=?", (did,)).fetchone()
    if not dep:
        return jsonify({"error": "Dependencia no encontrada"}), 404
        
    db.execute(
        "UPDATE dependencias SET liberacion_forzada=1, forzada_por=?, forzada_fecha=?, forzada_nota=? WHERE id=?",
        (admin_user, fecha_hoy, nota, did)
    )
    db.execute(
        "INSERT INTO avisos (fecha, proveedor, usuario, tipo, detalle, visto) VALUES (?, ?, ?, ?, ?, 0)",
        (datetime.datetime.now().isoformat(timespec="seconds"), "Admin", admin_user, "dependencia",
         f"Liberación anticipada forzada de {dep['tipo_sucesor']} en {dep['area']}: {nota}")
    )
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/dependencias/<int:did>/deshacer_forzar", methods=["POST"])
@requiere_gestor
def api_dependencias_deshacer_forzar(did):
    db = get_db()
    db.execute("UPDATE dependencias SET liberacion_forzada=0, forzada_por=NULL, forzada_fecha=NULL, forzada_nota=NULL WHERE id=?", (did,))
    db.commit()
    return jsonify({"ok": True})


# ----------------------------------------------------------------------------
# API — Reportes para Departamentos Internos (PDF / Excel)
# ----------------------------------------------------------------------------
@app.route("/api/hoja_departamento.pdf")
def hoja_departamento_pdf():
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    depto = request.args.get("departamento", "")
    titulo_txt = request.args.get("titulo", "")
    
    db = get_db()
    cond = ["(mundo='interno' OR departamento IS NOT NULL)"]
    args = []
    if depto:
        cond.append("departamento = ?"); args.append(depto)
    for campo in ("bloque", "area", "estatus"):
        v = request.args.get(campo)
        if v:
            cond.append(f"{campo} = ?"); args.append(v)
            
    q = f"SELECT * FROM actividades WHERE {' AND '.join(cond)} ORDER BY bloque, area, id"
    rows = db.execute(q, args).fetchall()
    
    hoy_d = datetime.date.today()
    out = os.path.join(BASE_DIR, "data", "hoja_departamento.pdf")
    doc = SimpleDocTemplate(out, pagesize=landscape(letter),
                            leftMargin=1*cm, rightMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    titulo = ParagraphStyle("t", parent=styles["Title"], fontSize=13, textColor=colors.HexColor("#3A2C52"))
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=7, leading=8)
    
    encabezado = titulo_txt or (f"Hoja de actualización de avance — Departamento: {depto}" if depto else "Reporte de actividades — Departamentos Internos HAP")
    elems = [Paragraph(encabezado, titulo),
             Paragraph(f"Fecha: {hoy_d.strftime('%d/%m/%Y')} · Unidad Quirúrgica Torre B · Uso interno HAP", styles["Normal"]),
             Spacer(1, 8)]
             
    data = [["Código", "Bloque / Área", "Tipo tarea", "Actividad / Partida", "Avance\nactual", "NUEVO\navance", "Fecha\ncompromiso", "Retra-\nsada", "Observaciones"]]
    filas_retraso = []
    
    for idx, r in enumerate(rows):
        fin = parse_date(r["f_fin"])
        retrasada = fin and (r["avance"] or 0) < 100 and (fin - hoy_d).days < 0
        bloque_area = f"{r['bloque'] or ''}\n{r['area'] or ''}"
        data.append([
            r["codigo"] or "",
            Paragraph(bloque_area, small),
            Paragraph(r["tipo_interno"] or r["tipo"] or "Instalación", small),
            Paragraph(r["partida"] or "", small),
            f"{r['avance'] or 0}%",
            "______",
            r["f_fin"] or "__________",
            "SÍ" if retrasada else "",
            "________________"
        ])
        if retrasada:
            filas_retraso.append(idx + 1)
            
    t = Table(data, colWidths=[2*cm, 4*cm, 2.8*cm, 6.2*cm, 1.5*cm, 1.6*cm, 2.2*cm, 1.3*cm, 4.5*cm], repeatRows=1)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3A2C52")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F7FB")]),
    ]
    for fr in filas_retraso:
        estilo.append(("BACKGROUND", (0, fr), (-1, fr), colors.HexColor("#F8D7DA")))
    t.setStyle(TableStyle(estilo))
    elems.append(t)
    doc.build(elems)
    safe = "".join(ch for ch in (depto or "departamento") if ch.isalnum() or ch in " _-")[:30].strip() or "depto"
    return send_file(out, as_attachment=True, download_name=f"HOJA_DEPTO_{safe}_{hoy()}.pdf")


@app.route("/api/hoja_departamento.xlsx")
def hoja_departamento_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    depto = request.args.get("departamento", "")
    db = get_db()
    cond = ["(mundo='interno' OR departamento IS NOT NULL)"]
    args = []
    if depto:
        cond.append("departamento = ?"); args.append(depto)
    for campo in ("bloque", "area", "estatus"):
        v = request.args.get(campo)
        if v:
            cond.append(f"{campo} = ?"); args.append(v)
            
    q = f"SELECT * FROM actividades WHERE {' AND '.join(cond)} ORDER BY bloque, area, id"
    rows = db.execute(q, args).fetchall()
    
    hoy_d = datetime.date.today()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Actualización Depto"
    encabezado = f"HOJA DE ACTUALIZACIÓN DE AVANCE — DEPARTAMENTO: {depto or 'INTERNO'}"
    ws["A1"] = encabezado
    ws["A1"].font = Font(bold=True, size=13, color="3A2C52")
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Fecha: {hoy_d.strftime('%d/%m/%Y')}   ·   Uso interno HAP — Anotar nuevo avance y fecha compromiso."
    ws["A2"].font = Font(size=9, italic=True, color="555555")
    ws.merge_cells("A2:I2")
    
    cols = ["Código", "Bloque", "Área", "Tipo Tarea", "Actividad / Partida", "Avance actual",
            "NUEVO avance %", "Fecha compromiso", "Observaciones"]
    hdr = 4
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=hdr, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3A2C52")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    i = hdr + 1
    for r in rows:
        vals = [r["codigo"], r["bloque"], r["area"], r["tipo_interno"] or r["tipo"] or "Instalación",
                r["partida"], f"{r['avance'] or 0}%", "", r["f_fin"] or "", ""]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = bd
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        i += 1
    anchos = [10, 18, 22, 16, 38, 12, 13, 16, 28]
    for j, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"
    safe = "".join(ch for ch in (depto or "departamento") if ch.isalnum() or ch in " _-")[:30].strip() or "depto"
    out = os.path.join(BASE_DIR, "data", "hoja_departamento.xlsx")
    wb.save(out)
    return send_file(out, as_attachment=True, download_name=f"HOJA_DEPTO_{safe}_{hoy()}.xlsx")


# ----------------------------------------------------------------------------
# API — Respaldos de base de datos
# ----------------------------------------------------------------------------
@app.route("/api/respaldo", methods=["POST"])
@requiere_admin
def api_respaldo():
    res = generar_respaldo_bd()
    if res:
        return jsonify({"ok": True, "archivo": os.path.basename(res)})
    return jsonify({"error": "No se pudo generar el respaldo"}), 500


@app.route("/api/respaldo/descargar", methods=["GET"])
@requiere_admin
def api_respaldo_descargar():
    if os.path.exists(DB_PATH):
        return send_file(DB_PATH, as_attachment=True, download_name=f"obra_backup_{datetime.date.today().isoformat()}.db")
    return jsonify({"error": "Base de datos no encontrada"}), 404


@app.route("/api/validacion/panel")
def validacion_page():
    return render_template("validacion.html")


@app.route("/usuarios")
def usuarios_page():
    return render_template("usuarios.html")


@app.route("/validacion")
def validacion_page_corta():
    return render_template("validacion.html")



def _sembrar_si_vacia():
    """Primera vez en un servidor nuevo: si el disco esta vacio, copia la
    base semilla con las 955 actividades. Asi no hay que subirla a mano."""
    semilla = os.path.join(BASE_DIR, "semilla", "obra_inicial.db")
    if not os.path.exists(semilla):
        return
    try:
        necesita = not os.path.exists(DB_PATH)
        if not necesita:
            con = sqlite3.connect(DB_PATH)
            try:
                n = con.execute("SELECT COUNT(*) FROM actividades").fetchone()[0]
                necesita = (n == 0)
            except sqlite3.Error:
                necesita = True
            con.close()
        if necesita:
            import shutil
            shutil.copy(semilla, DB_PATH)
            print("  Base inicial sembrada desde semilla/obra_inicial.db")
    except Exception as e:
        print("  (no se pudo sembrar la base inicial:", e, ")")


# --- Arranque cuando corre en internet (gunicorn no ejecuta el bloque de abajo) ---
_ARRANQUE_MODULO = True
try:
    _sembrar_si_vacia()
    init_db()
    _foto_automatica()
except Exception as _e:
    print("Aviso al iniciar la base:", _e)


if __name__ == "__main__":
    print("=" * 60)
    print("  PLATAFORMA DE CONTROL DE OBRA — HAP")
    print("  Base de datos:", DB_PATH)
    print("  Abre en tu navegador:  http://localhost:5000")
    print("=" * 60)
    puerto = int(os.environ.get("PORT", 5000))
    host = "0.0.0.0" if os.environ.get("EN_INTERNET") else "127.0.0.1"
    app.run(host=host, port=puerto, debug=False)
