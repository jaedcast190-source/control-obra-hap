#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importa el maestro de Excel a la base de datos local.

Uso:
    python importar.py "ruta/al/BORRADOR_MAESTRO_REVISION.xlsx"

Si no pasas ruta, busca automáticamente un .xlsx en esta carpeta.
Puedes correrlo cuantas veces quieras: te preguntará si reemplazar
lo que ya haya cargado.
"""

import os
import sys
import sqlite3
import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "obra.db")


def normaliza_avance(v):
    if v is None:
        return 0
    try:
        f = float(v)
        if f <= 1.0:  # por si viene como 0.25 = 25%
            f *= 100
        return int(round(f))
    except (ValueError, TypeError):
        return 0


def main():
    try:
        import openpyxl
    except ImportError:
        print("Falta openpyxl. Instálalo con:  pip install openpyxl")
        sys.exit(1)

    # localizar archivo
    if len(sys.argv) > 1:
        xlsx = sys.argv[1]
    else:
        candidatos = [f for f in os.listdir(BASE_DIR) if f.lower().endswith(".xlsx")]
        if not candidatos:
            print("No encontré ningún .xlsx. Pásame la ruta:")
            print('   python importar.py "C:\\ruta\\BORRADOR_MAESTRO_REVISION.xlsx"')
            sys.exit(1)
        xlsx = os.path.join(BASE_DIR, candidatos[0])
        print("Usando archivo:", candidatos[0])

    if not os.path.exists(xlsx):
        print("No existe el archivo:", xlsx)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["MAESTRO"] if "MAESTRO" in wb.sheetnames else wb.active

    # detectar fila de encabezados (busca 'ID' en col 1)
    header_row = None
    for r in range(1, 10):
        if str(ws.cell(row=r, column=1).value).strip().upper() == "ID":
            header_row = r
            break
    if header_row is None:
        header_row = 4
    start = header_row + 1

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.executescript(open(os.path.join(BASE_DIR, "schema.sql")).read()
                      if os.path.exists(os.path.join(BASE_DIR, "schema.sql")) else "")

    # crear tablas si no existen (mismo esquema que app.py)
    con.executescript(
        """
    CREATE TABLE IF NOT EXISTS actividades (
        id INTEGER PRIMARY KEY AUTOINCREMENT, codigo TEXT, bloque TEXT, area TEXT,
        giro TEXT, proveedor TEXT, partida TEXT, tipo TEXT, aplica TEXT DEFAULT 'SÍ',
        avance INTEGER DEFAULT 0, f_inicio TEXT, f_fin TEXT, duracion_dias INTEGER,
        estatus TEXT DEFAULT 'Pendiente', depende_de INTEGER, notas TEXT, actualizado TEXT);
    CREATE TABLE IF NOT EXISTS historial (
        id INTEGER PRIMARY KEY AUTOINCREMENT, actividad_id INTEGER, campo TEXT,
        valor_antes TEXT, valor_despues TEXT, fecha TEXT);
    CREATE TABLE IF NOT EXISTS config (clave TEXT PRIMARY KEY, valor TEXT);
    """
    )

    cur = con.execute("SELECT COUNT(*) FROM actividades")
    existentes = cur.fetchone()[0]
    if existentes > 0:
        resp = input(f"Ya hay {existentes} actividades cargadas. ¿Reemplazar todas? (s/n): ")
        if resp.strip().lower() == "s":
            con.execute("DELETE FROM actividades")
            con.execute("DELETE FROM historial")
            con.commit()
            print("Datos anteriores borrados.")
        else:
            print("Importación cancelada. No se tocó nada.")
            con.close()
            return

    n = 0
    for r in range(start, ws.max_row + 1):
        codigo = ws.cell(row=r, column=1).value
        if not codigo:
            continue
        bloque = ws.cell(row=r, column=2).value
        area = ws.cell(row=r, column=3).value
        giro = ws.cell(row=r, column=4).value
        proveedor = ws.cell(row=r, column=5).value
        partida = ws.cell(row=r, column=6).value
        tipo = ws.cell(row=r, column=7).value
        aplica = ws.cell(row=r, column=8).value or "SÍ"
        avance = normaliza_avance(ws.cell(row=r, column=9).value)
        f_fab = ws.cell(row=r, column=10).value
        f_inst = ws.cell(row=r, column=11).value
        f_ent = ws.cell(row=r, column=12).value
        estatus = ws.cell(row=r, column=13).value or "Pendiente"
        notas = ws.cell(row=r, column=14).value

        # F. Inicio = fabricación (si hay) o instalación; F. Fin = entrega o instalación
        def fmt(v):
            if v is None:
                return None
            if isinstance(v, (datetime.date, datetime.datetime)):
                return v.strftime("%Y-%m-%d")
            return str(v)

        f_inicio = fmt(f_fab) or fmt(f_inst)
        f_fin = fmt(f_ent) or fmt(f_inst)

        con.execute(
            """INSERT INTO actividades
            (codigo,bloque,area,giro,proveedor,partida,tipo,aplica,avance,
             f_inicio,f_fin,estatus,notas,actualizado)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (str(codigo), bloque, area, giro, proveedor, partida, tipo,
             str(aplica), avance, f_inicio, f_fin, estatus, notas,
             datetime.datetime.now().isoformat(timespec="seconds")),
        )
        n += 1

    con.commit()
    con.close()
    print(f"\n✓ Importadas {n} actividades a la base de datos.")
    print(f"  Base de datos: {DB_PATH}")
    print(f"\nAhora corre:  python app.py")


if __name__ == "__main__":
    main()
